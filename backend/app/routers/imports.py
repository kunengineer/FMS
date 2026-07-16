from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.core.database import get_db
from app.core.dependencies import PermissionChecker
from app.models import (
    Vehicle, Operator, VehicleType, Role, Shift, OperationLog,
    FailureLog, FailureCategory, RepairLog, ChecklistItem, ChecklistResult,
    ConditionStatus, VehicleStatus, SeverityLevel, RepairStatus
)
from app.core.security import get_password_hash
import openpyxl
from io import BytesIO
from datetime import datetime, date, time, timedelta
from typing import List, Optional
from decimal import Decimal
import uuid
import re
import unicodedata

def parse_date(val) -> Optional[date]:
    if isinstance(val, datetime):
        return val.date()
    elif isinstance(val, date):
        return val
    if not val:
        return None
    val_str = str(val).strip()
    # Take the date part only (e.g. "2026-07-15 00:00:00" -> "2026-07-15")
    val_str = val_str.split()[0]
    
    # Try common formats
    for fmt in (
        "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d",
        "%d-%m-%Y", "%d.%m.%Y", "%Y.%m.%d"
    ):
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            pass
            
    # Regex fallback for dd/mm/yyyy or dd-mm-yyyy or dd.mm.yyyy
    m = re.search(r'(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})', val_str)
    if m:
        try:
            d = int(m.group(1))
            m_val = int(m.group(2))
            y = int(m.group(3))
            return date(y, m_val, d)
        except ValueError:
            pass
            
    # Regex fallback for yyyy/mm/dd or yyyy-mm-dd or yyyy.mm.dd
    m_rev = re.search(r'(\d{4})[/\-\.](\d{1,2})[/\-\.](\d{1,2})', val_str)
    if m_rev:
        try:
            y = int(m_rev.group(1))
            m_val = int(m_rev.group(2))
            d = int(m_rev.group(3))
            return date(y, m_val, d)
        except ValueError:
            pass
            
    return None

router = APIRouter(prefix="/imports", tags=["imports"])

class ExcelWrapper:
    def __init__(self, contents: bytes, filename: str):
        self.filename = filename
        is_xls = filename.lower().endswith('.xls')
        if not is_xls:
            if len(contents) > 4 and contents[:4] != b'PK\x03\x04':
                is_xls = True
                
        if is_xls:
            try:
                import xlrd
                self.is_xls = True
                self.wb = xlrd.open_workbook(file_contents=contents)
                self.sheet_names = self.wb.sheet_names()
            except Exception as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Không thể mở tập tin Excel (.xls): {str(e)}. Đảm bảo thư viện 'xlrd' được cài đặt."
                )
        else:
            self.is_xls = False
            try:
                self.wb = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
                self.sheet_names = self.wb.sheetnames
            except Exception as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Không thể mở tập tin Excel (.xlsx): {str(e)}"
                )

    def get_sheet(self, name: str):
        if self.is_xls:
            return self.wb.sheet_by_name(name)
        else:
            return self.wb[name]

    def get_cell_value(self, sheet, row: int, col: int):
        if self.is_xls:
            if row - 1 < sheet.nrows and col - 1 < sheet.ncols:
                val = sheet.cell_value(row - 1, col - 1)
                if val == "":
                    return None
                return val
            return None
        else:
            return sheet.cell(row=row, column=col).value

    def get_max_row(self, sheet) -> int:
        if self.is_xls:
            return sheet.nrows
        else:
            return sheet.max_row

    def get_max_column(self, sheet) -> int:
        if self.is_xls:
            return sheet.ncols
        else:
            return sheet.max_column

def extract_date_from_sheet(sheet, wrapper, sheet_name: str = "", filename: str = "") -> Optional[date]:
    # 1. First try to parse week number from sheet name and calculate the Friday of that week
    if sheet_name:
        # Check YYYY-Www pattern first (e.g., 2026-W12 or 2026-W02)
        m_iso = re.search(r'(\d{4})-W(\d+)', sheet_name, re.IGNORECASE)
        if m_iso:
            try:
                year = int(m_iso.group(1))
                week_num = int(m_iso.group(2))
                return date.fromisocalendar(year, week_num, 5) # Friday of that week
            except Exception:
                pass

        m_week = re.search(r'(?:tuần|tuan)\s*(\d+)', sheet_name, re.IGNORECASE)
        if m_week:
            try:
                week_num = int(m_week.group(1))
                year = 2026  # Default fallback
                if filename:
                    m_year = re.search(r'20\d{2}', filename)
                    if m_year:
                        year = int(m_year.group(0))
                # ISO week 1 starts on the Monday of the week containing the first Thursday of the year
                jan4 = date(year, 1, 4)
                week1_monday = jan4 - timedelta(days=jan4.weekday())
                week1_friday = week1_monday + timedelta(days=4)
                return week1_friday + timedelta(weeks=week_num - 1)
            except Exception:
                pass

    # 2. Fallback to scanning first 5 rows for text matching: "Ngày DD tháng MM năm YYYY"
    date_regex = re.compile(r'ng[àa]y\s+(\d+)\s+th[áa]ng\s+(\d+)\s+n[ăa]m\s+(\d+)', re.IGNORECASE)
    max_rows = wrapper.get_max_row(sheet)
    max_cols = wrapper.get_max_column(sheet)
    for r in range(1, min(6, max_rows + 1)):
        for c in range(1, min(max_cols + 1, 15)):
            val = wrapper.get_cell_value(sheet, r, c)
            if isinstance(val, (datetime, date)):
                return val.date() if isinstance(val, datetime) else val
            if isinstance(val, str):
                m = date_regex.search(val)
                if m:
                    try:
                        d = int(m.group(1))
                        m_val = int(m.group(2))
                        y = int(m.group(3))
                        return date(y, m_val, d)
                    except ValueError:
                        pass
                
                # Fallback to parsing standard date formats (like 15/07/2026)
                parsed = parse_date(val)
                if parsed:
                    return parsed
    return None

def find_header_row_and_cols(sheet, wrapper):
    max_rows = wrapper.get_max_row(sheet)
    max_cols = wrapper.get_max_column(sheet)
    
    header_row_idx = None
    cols = {
        "stt": None,
        "name": None,
        "code": None,
        "hours": None,
        "hourmeter_start": None,
        "hourmeter_end": None,
        "active": None,
        "inactive": None,
        "condition": None,
        "status_text": None,
        "downtime": None,
        "repair_count": None,
        "repair_done": None,
        "unresolved": None,
        "notes": None,
        "desc_detail": None,
        "downtime_desc": None,
        "backlog_date": None,
        "suggestion": None
    }
    
    for r in range(1, min(16, max_rows + 1)):
        row_vals = [wrapper.get_cell_value(sheet, r, c) for c in range(1, max_cols + 1)]
        row_strs = [re.sub(r'\s+', ' ', str(val)).strip().lower() for val in row_vals if val is not None]
        
        non_empty_count = sum(1 for val in row_vals if val is not None and str(val).strip() != "")
        if non_empty_count < 5:
            continue
            
        has_stt = any("stt" in s for s in row_strs)
        has_name = any("thiết bị" in s or "phương tiện" in s or "tên" in s for s in row_strs)
        has_cond = any("hiện trạng" in s or "hư hỏng" in s or "tình trạng" in s or "hạng mục" in s for s in row_strs)
        
        if (has_stt and has_name) or (has_name and has_cond):
            header_row_idx = r
            for col_idx in range(1, max_cols + 1):
                val = wrapper.get_cell_value(sheet, r, col_idx)
                if not val:
                    continue
                val_str = re.sub(r'\s+', ' ', str(val)).strip().lower()
                
                if "stt" in val_str:
                    cols["stt"] = col_idx
                elif "thiết bị" in val_str or "phương tiện" in val_str:
                    if cols["name"] is None or "tên" in val_str:
                        cols["name"] = col_idx
                elif "mã" in val_str or "mã" in val_str:
                    cols["code"] = col_idx
                elif "giờ đầu" in val_str or "giờ đầu" in val_str or "giờ máy đầu" in val_str or "giờ chạy đầu" in val_str:
                    cols["hourmeter_start"] = col_idx
                elif "giờ cuối" in val_str or "giờ cuối" in val_str or "giờ máy cuối" in val_str or "giờ chạy cuối" in val_str:
                    cols["hourmeter_end"] = col_idx
                elif "giờ hoạt động" in val_str or "giờ chạy" in val_str or "giờ máy" in val_str or val_str == "giờ hđ" or val_str == "giờ hd":
                    cols["hours"] = col_idx
                elif val_str == "hoạt động" or val_str == "hoạt đông":
                    cols["active"] = col_idx
                elif "không hoạt động" in val_str or "không hoạt đông" in val_str:
                    cols["inactive"] = col_idx
                elif "chi tiết hư hỏng" in val_str or "chi tiết hạng mục" in val_str or "mô tả chi tiết" in val_str:
                    cols["desc_detail"] = col_idx
                elif "hạng mục" in val_str or "hiện trạng" in val_str or "nội dung hư hỏng" in val_str:
                    cols["condition"] = col_idx
                elif "tình trạng" in val_str or "trạng thái" in val_str:
                    cols["status_text"] = col_idx
                elif "mô tả tg sc" in val_str or "mô tả thời gian sc" in val_str or "mô tả thời gian sữa chữa" in val_str:
                    cols["downtime_desc"] = col_idx
                elif "giờ sc" in val_str or "thời gian dừng" in val_str or "thời gian hư" in val_str or ("thời gian" in val_str and ("hư" in val_str or "kiểm tra" in val_str or "sửa chữa" in val_str)):
                    cols["downtime"] = col_idx
                elif "số lần sửa" in val_str or "số lấn sửa" in val_str or "sl sc" in val_str:
                    cols["repair_count"] = col_idx
                elif "hoàn thành" in val_str or "xong" in val_str or "khắc phục" in val_str:
                    cols["repair_done"] = col_idx
                elif "tồn đọng từ ngày" in val_str or "tồn đọng từ" in val_str:
                    cols["backlog_date"] = col_idx
                elif "chưa xử lý" in val_str or "chưa triệt để" in val_str or "tồn đọng" in val_str:
                    cols["unresolved"] = col_idx
                elif "đề nghị" in val_str or "kiến nghị" in val_str:
                    cols["suggestion"] = col_idx
                elif "ghi chú" in val_str or "ghi chú" in val_str:
                    cols["notes"] = col_idx
            break
            
    return header_row_idx, cols

def generate_operator_id(name: str, prefix: str = "OP") -> str:
    nfkd_form = unicodedata.normalize('NFKD', name)
    only_ascii = "".join([c for c in nfkd_form if not unicodedata.combining(c)])
    clean = re.sub(r'[^a-zA-Z0-9]', '', only_ascii).upper()
    base = f"{prefix}_{clean[:14]}"
    return base[:20]

def get_or_create_operator(db: Session, full_name: str, role_name: str) -> Operator:
    full_name = full_name.strip()
    if not full_name:
        full_name = "Chưa biết"
    
    # 1. Match exact name in active personnel database
    op = db.query(Operator).filter(func.lower(Operator.full_name) == full_name.lower(), Operator.active == True).first()
    if op:
        return op
        
    # 2. Allow active default users like default mechanic or driver
    if full_name in ("Thợ cơ điện mặc định", "Tài xế mặc định"):
        op_any = db.query(Operator).filter(func.lower(Operator.full_name) == full_name.lower()).first()
        if op_any:
            return op_any

    # 3. Fallback to single "UNKNOWN" operator ("Chưa biết")
    unknown_op = db.query(Operator).filter(Operator.operator_id == "UNKNOWN").first()
    if not unknown_op:
        role = db.query(Role).filter(Role.role_name == "NGƯỜI VẬN HÀNH").first()
        if not role:
            role = db.query(Role).first()
        role_id = role.role_id if role else 1
        
        unknown_op = Operator(
            operator_id="UNKNOWN",
            full_name="Chưa biết",
            department="Chưa xác định",
            role_id=role_id,
            phone=None,
            password_hash=get_password_hash("123456"),
            active=True
        )
        db.add(unknown_op)
        db.commit()
        db.refresh(unknown_op)
        
    return unknown_op

def normalize_vehicle_code(code: str) -> str:
    # Remove all spaces and convert to uppercase
    code = re.sub(r'\s+', '', code).upper()
    # Strip leading zeros after a dash, e.g., KM25-01 -> KM25-1
    code = re.sub(r'-0+(\d+)', r'-\1', code)
    return code

def get_or_create_vehicle(db: Session, code: str, type_name: str) -> Vehicle:
    code = code.strip()
    code_clean = normalize_vehicle_code(code)
    
    # 1. Try exact case/space-insensitive match on the clean code first (active only)
    veh = db.query(Vehicle).filter(
        Vehicle.active == True,
        func.replace(func.upper(Vehicle.vehicle_code), ' ', '') == code_clean
    ).first()
    if veh:
        return veh
        
    # 2. Try matching normalized codes in DB (active only)
    all_vehs = db.query(Vehicle).filter(Vehicle.active == True).all()
    for v in all_vehs:
        if normalize_vehicle_code(v.vehicle_code) == code_clean:
            return v
            
    # 3. Alphanumeric fallback match (active only)
    code_alphanum = re.sub(r'[^A-Z0-9]', '', code_clean)
    for v in all_vehs:
        v_alphanum = re.sub(r'[^A-Z0-9]', '', normalize_vehicle_code(v.vehicle_code))
        if v_alphanum == code_alphanum:
            return v
    
    vtype = db.query(VehicleType).filter(func.lower(VehicleType.type_name) == type_name.lower()).first()
    if not vtype:
        vtype = VehicleType(type_name=type_name)
        db.add(vtype)
        db.commit()
        db.refresh(vtype)
        
    status_code = "active"
    status_exists = db.query(VehicleStatus).filter_by(status_code=status_code).first()
    if not status_exists:
        db.add(VehicleStatus(status_code=status_code, status_label="Hoạt động"))
        db.commit()

    new_veh = Vehicle(
        vehicle_id=uuid.uuid4(),
        vehicle_code=code,
        vehicle_name=f"{type_name} {code}",
        vehicle_type_id=vtype.vehicle_type_id,
        status="active",
        current_hourmeter=0.0,
        last_maintenance_hourmeter=0.0,
        active=True
    )
    db.add(new_veh)
    db.commit()
    db.refresh(new_veh)
    return new_veh

def get_failure_category(db: Session, desc: str) -> int:
    desc_lower = desc.lower()
    cat_keyword_map = {
        "động cơ": "Hỏng động cơ",
        "máy": "Hỏng động cơ",
        "phanh": "Mất phanh / Hỏng phanh",
        "thắng": "Mất phanh / Hỏng phanh",
        "ống": "Xì bể ống thủy lực",
        "thủy lực": "Xì bể ống thủy lực",
        "lốp": "Bể lốp / Nổ lốp",
        "bánh": "Bể lốp / Nổ lốp",
        "cáp": "Đứt cáp cẩu",
        "ắc quy": "Hư bình ắc quy / Không đề được",
        "đề": "Hư bình ắc quy / Không đề được",
        "bình": "Hư bình ắc quy / Không đề được"
    }
    
    target_name = "Hư hỏng chung"
    for kw, cat_name in cat_keyword_map.items():
        if kw in desc_lower:
            target_name = cat_name
            break
            
    cat = db.query(FailureCategory).filter(FailureCategory.category_name == target_name).first()
    if not cat:
        # Severity Levels lookup
        sev_exists = db.query(SeverityLevel).filter_by(severity_code="light").first()
        if not sev_exists:
            db.add(SeverityLevel(severity_code="light", severity_label="Nhẹ"))
            db.commit()
            
        cat = FailureCategory(category_name=target_name, severity_default="light")
        db.add(cat)
        db.commit()
        db.refresh(cat)
        
    return cat.category_id

def parse_time(val) -> Optional[time]:
    if isinstance(val, time):
        return val.replace(microsecond=0)
    elif isinstance(val, datetime):
        return val.time().replace(microsecond=0)
    elif isinstance(val, str):
        for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M%p"):
            try:
                return datetime.strptime(val.strip(), fmt).time().replace(microsecond=0)
            except ValueError:
                pass
    return None

def get_shift_from_time(db: Session, t: time) -> int:
    shifts = db.query(Shift).all()
    if not shifts:
        # Create standard shift
        s1 = Shift(shift_id=1, shift_name="Ca Ngày", start_time=time(6, 0), end_time=time(18, 0))
        db.add(s1)
        db.commit()
        return 1
        
    for s in shifts:
        start = s.start_time
        end = s.end_time
        if start <= end:
            if start <= t <= end:
                return s.shift_id
        else:
            if t >= start or t <= end:
                return s.shift_id
    return shifts[0].shift_id

@router.post("/activity")
async def import_activity_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Operator = Depends(PermissionChecker(["reports:view"]))
):
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không thể mở tập tin Excel: {str(e)}")

    # Ensure required tables have statuses
    for code, label in {"ok": "Bình thường", "broken": "Có hư hỏng"}.items():
        if not db.query(ConditionStatus).filter_by(status_code=code).first():
            db.add(ConditionStatus(status_code=code, status_label=label))
    for code, label in {"pending": "Chờ sửa chữa", "in_progress": "Đang sửa chữa", "done": "Đã hoàn thành"}.items():
        if not db.query(RepairStatus).filter_by(status_code=code).first():
            db.add(RepairStatus(status_code=code, status_label=label))
    db.commit()

    stats = {
        "sheets_processed": 0,
        "rows_processed": 0,
        "vehicles_created": 0,
        "operators_created": 0,
        "operation_logs_created": 0,
        "failure_logs_created": 0,
        "repair_logs_created": 0,
        "errors": []
    }

    initial_vehicle_count = db.query(Vehicle).count()
    initial_operator_count = db.query(Operator).count()

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Determine vehicle type based on sheet name
        type_name = "Thiết bị"
        if "cẩu" in sheet_name.lower() or "ray" in sheet_name.lower():
            type_name = "Cần cẩu"
        elif "cuốc" in sheet_name.lower() or "đào" in sheet_name.lower():
            type_name = "Xe cuốc"
        elif "nâng" in sheet_name.lower():
            type_name = "Xe nâng"
        elif "gàu" in sheet_name.lower():
            type_name = "Gàu hoa thị"
            
        # Parse Row 2 and Row 3 to map columns
        if sheet.max_row < 3:
            continue
            
        vehicle_cols = {}
        header_cols = {
            "date": None,
            "status": None,
            "from": None,
            "to": None,
            "duration": None,
            "notes": None,
            "performers": None,
            "gm": None,
            "failure_name": None
        }
        
        # Scan columns
        for col_idx in range(1, sheet.max_column + 1):
            r2_val = sheet.cell(row=2, column=col_idx).value
            r3_val = sheet.cell(row=3, column=col_idx).value
            
            # Map vehicle columns
            if r3_val and (r2_val is None or "phương tiện" in str(r2_val).strip().lower()):
                vehicle_cols[col_idx] = str(r3_val).strip().replace('\n', ' ')
                
            # Map other columns based on Row 2
            if r2_val:
                val_str = str(r2_val).strip().lower()
                if "thời gian" in val_str and "hư" not in val_str:
                    header_cols["date"] = col_idx
                elif "hư hỏng" in val_str or "công việc" in val_str or "nội dung" in val_str:
                    header_cols["status"] = col_idx
                elif val_str == "từ":
                    header_cols["from"] = col_idx
                elif val_str == "đến":
                    header_cols["to"] = col_idx
                elif "thời gian" in val_str and "hư" in val_str:
                    header_cols["duration"] = col_idx
                elif "ghi chú" in val_str:
                    header_cols["notes"] = col_idx
                elif "người" in val_str and "thực hiện" in val_str:
                    header_cols["performers"] = col_idx
                elif val_str == "gm" or "chỉ số" in val_str or "giờ máy" in val_str:
                    header_cols["gm"] = col_idx
                elif "sự cố" in val_str:
                    header_cols["failure_name"] = col_idx

        # If we couldn't find date column, skip sheet
        if not header_cols["date"]:
            continue
            
        stats["sheets_processed"] += 1
        
        # Start reading rows from row 4 onwards
        last_valid_date = None
        last_valid_vehicle = None
        
        for r in range(4, sheet.max_row + 1):
            # Check date value
            date_cell = sheet.cell(row=r, column=header_cols["date"]).value
            
            if date_cell is not None:
                work_date = parse_date(date_cell)
                if not work_date:
                    stats["errors"].append({
                        "sheet": sheet.title,
                        "row": r,
                        "message": f"Ngày '{date_cell}' không đúng định dạng ngày tháng, bỏ qua dòng."
                    })
                    continue
                else:
                    last_valid_date = work_date
            else:
                # If date is empty, check if we have other fields on this row
                # If everything in the row is empty, just skip it silently
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
                if not any(v is not None for v in row_vals):
                    continue
                # Otherwise, carry forward the date
                if last_valid_date:
                    work_date = last_valid_date
                else:
                    continue
                    
            # Find which vehicle has 'x' in this row
            target_vehicle_code = None
            for col_idx, vcode in vehicle_cols.items():
                cell_val = sheet.cell(row=r, column=col_idx).value
                if cell_val and str(cell_val).strip().lower() == 'x':
                    target_vehicle_code = vcode
                    break
                    
            if target_vehicle_code:
                last_valid_vehicle = target_vehicle_code
            else:
                if last_valid_vehicle:
                    target_vehicle_code = last_valid_vehicle
                else:
                    stats["errors"].append({
                        "sheet": sheet.title,
                        "row": r,
                        "message": "Không tìm thấy phương tiện nào được đánh dấu 'x' vận hành, bỏ qua."
                    })
                    continue
                    
            stats["rows_processed"] += 1
                
            # Get or create vehicle
            vehicle = get_or_create_vehicle(db, target_vehicle_code, type_name)
            
            # Parse parameters
            status_text = ""
            if header_cols["status"]:
                status_cell = sheet.cell(row=r, column=header_cols["status"]).value
                status_text = str(status_cell).strip() if status_cell else ""
                
            from_time = time(8, 0)
            if header_cols["from"]:
                t_val = parse_time(sheet.cell(row=r, column=header_cols["from"]).value)
                if t_val:
                    from_time = t_val
                    
            to_time = time(16, 0)
            if header_cols["to"]:
                t_val = parse_time(sheet.cell(row=r, column=header_cols["to"]).value)
                if t_val:
                    to_time = t_val
                    
            notes = ""
            if header_cols["notes"]:
                notes_cell = sheet.cell(row=r, column=header_cols["notes"]).value
                notes = str(notes_cell).strip() if notes_cell else ""
                
            gm_value = 0.0
            if header_cols["gm"]:
                gm_cell = sheet.cell(row=r, column=header_cols["gm"]).value
                try:
                    gm_value = float(gm_cell) if gm_cell is not None else 0.0
                except (ValueError, TypeError):
                    pass
            # Update vehicle current hourmeter if gm_value is higher
            if gm_value > float(vehicle.current_hourmeter):
                vehicle.current_hourmeter = gm_value
                db.commit()

            performers_text = ""
            if header_cols["performers"]:
                perf_cell = sheet.cell(row=r, column=header_cols["performers"]).value
                performers_text = str(perf_cell).strip() if perf_cell else ""
                
            # Determine operators/performers
            performer_ops = []
            if performers_text:
                names = re.split(r'[,;\-\n/]| và ', performers_text, flags=re.IGNORECASE)
                for name in names:
                    name = name.strip()
                    if name and name.lower() not in ("bd", "xck", "tổ bảo trì", "bảo dưỡng"):
                        mech = get_or_create_operator(db, name, "THỢ SỬA CHỮA")
                        performer_ops.append(mech)
                        
            # Determine main operator (default OP01, or first helper if any)
            main_op = get_or_create_operator(db, "Nguyễn Văn Vận Hành", "NGƯỜI VẬN HÀNH")
            
            # Determine Shift
            shift_id = get_shift_from_time(db, from_time)
            
            # Check if OperationLog already exists
            # Check if OperationLog already exists
            op_log = db.query(OperationLog).filter_by(
                vehicle_id=vehicle.vehicle_id,
                work_date=work_date,
                shift_id=shift_id
            ).first()
            
            # Read failure name if failure_name column is mapped
            failure_name = ""
            if header_cols.get("failure_name"):
                fn_cell = sheet.cell(row=r, column=header_cols["failure_name"]).value
                failure_name = str(fn_cell).strip() if fn_cell else ""

            is_failure = (
                "hỏng" in status_text.lower() or 
                "sự cố" in status_text.lower() or 
                "chảy nhớt" in notes.lower() or 
                "hư" in status_text.lower() or 
                "hư" in notes.lower() or
                bool(failure_name)
            )
            
            if is_failure:
                if failure_name:
                    failure_desc = failure_name
                else:
                    failure_desc = f"{status_text}: {notes}" if status_text and notes else (status_text or notes or "Hư hỏng không rõ chi tiết")
                
                # Check duplicate failure log on same vehicle and date
                exist_fail = db.query(FailureLog).filter(
                    FailureLog.vehicle_id == vehicle.vehicle_id,
                    FailureLog.description == failure_desc,
                    func.date(FailureLog.failure_time) == work_date
                ).first()
                if exist_fail:
                    continue # Skip duplicate row entirely
            elif op_log:
                continue # Ca already exists and this is a normal running row, skip
                
            if not op_log:
                op_log = OperationLog(
                    vehicle_id=vehicle.vehicle_id,
                    operator_id=main_op.operator_id,
                    shift_id=shift_id,
                    work_date=work_date,
                    start_hour=from_time,
                    end_hour=to_time,
                    hourmeter_start=gm_value,
                    hourmeter_end=gm_value,
                    condition_before_shift="broken" if is_failure else "ok",
                    is_safety_confirmed=True,
                    idempotency_key=uuid.uuid4(),
                    notes=notes or status_text
                )
                db.add(op_log)
                db.commit()
                db.refresh(op_log)
                stats["operation_logs_created"] += 1
                
            if is_failure:
                f_time = datetime.combine(work_date, from_time)
                category_id = get_failure_category(db, failure_desc)
                fail_log = FailureLog(
                    operation_id=op_log.operation_id,
                    vehicle_id=vehicle.vehicle_id,
                    category_id=category_id,
                    description=failure_desc,
                    failure_time=f_time,
                    severity="dangerous" if any(k in failure_desc.lower() for k in ("nguy hiểm", "mất phanh", "hỏng phanh", "cháy", "gãy", "đứt cáp")) else ("heavy" if "nặng" in failure_desc.lower() or "chảy nhớt" in failure_desc.lower() else "light"),
                    phase="during_shift",
                    is_repaired=True,
                    created_by=main_op.operator_id
                )
                db.add(fail_log)
                db.commit()
                db.refresh(fail_log)
                stats["failure_logs_created"] += 1
                
                # No Repair Log is created on import as per requirements

    final_vehicle_count = db.query(Vehicle).count()
    final_operator_count = db.query(Operator).count()
    
    stats["vehicles_created"] = final_vehicle_count - initial_vehicle_count
    stats["operators_created"] = final_operator_count - initial_operator_count

    return {
        "message": "Nhập báo cáo hoạt động phương tiện thành công",
        "statistics": stats
    }

@router.post("/checklist")
async def import_checklist_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Operator = Depends(PermissionChecker(["reports:view"]))
):
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không thể mở tập tin Excel: {str(e)}")

    if "Câu trả lời biểu mẫu 1" not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail="Tập tin checklist phải chứa trang bảng tính 'Câu trả lời biểu mẫu 1'"
        )
        
    sheet = wb["Câu trả lời biểu mẫu 1"]
    
    # Check headers
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    
    required_headers = ['dấu thời gian', 'họ tên người vận hành', 'tên phương tiện', 'số giờ hoạt động']
    header_indices = {}
    
    for req in required_headers:
        idx = None
        for i, h in enumerate(headers):
            if h and req in str(h).strip().lower():
                idx = i + 1
                break
        if not idx:
            raise HTTPException(
                status_code=400,
                detail=f"Thiếu cột bắt buộc trong file checklist: '{req}'"
            )
        header_indices[req] = idx

    # Map optional failure notes and safety confirmation columns
    special_columns = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        h_str = str(h).strip().lower()
        if 'đảm bảo an toàn' in h_str:
            special_columns['safety_confirmed'] = i + 1
        elif 'hư hỏng trong ca' in h_str or 'ghi chú hư hỏng' in h_str:
            special_columns['failure_note'] = i + 1

    # Every column starting from index 6 represents checklist items
    checklist_item_cols = {}
    blacklist_keywords = [
        "điểm số", "tổng điểm", "email", "địa chỉ", "phương tiện", "người vận hành", "giờ hoạt động", "stt", "cột",
        "cần cẩu đảm bảo an toàn để bắt đầu làm việc"
    ]
    for i, h in enumerate(headers):
        if h and i + 1 not in header_indices.values():
            h_str = str(h).strip()
            h_lower = h_str.lower()
            is_blacklisted = any(kw in h_lower for kw in blacklist_keywords)
            if not is_blacklisted:
                checklist_item_cols[i + 1] = h_str

    stats = {
        "rows_processed": 0,
        "vehicles_created": 0,
        "operators_created": 0,
        "checklist_items_created": 0,
        "operation_logs_created": 0,
        "checklist_results_created": 0,
        "failure_logs_created": 0,
        "repair_logs_created": 0,
        "errors": []
    }

    initial_vehicle_count = db.query(Vehicle).count()
    initial_operator_count = db.query(Operator).count()
    initial_checklist_count = db.query(ChecklistItem).count()

    # Pre-fetch or create ChecklistItems
    checklist_item_ids = {}
    for col_idx, item_name in checklist_item_cols.items():
        chk = db.query(ChecklistItem).filter(func.lower(ChecklistItem.item_name) == item_name.lower()).first()
        if not chk:
            chk = ChecklistItem(item_name=item_name, active=True, severity="light")
            db.add(chk)
            db.commit()
            db.refresh(chk)
        checklist_item_ids[col_idx] = chk.checklist_id

    # Read data rows (Row 2 onwards)
    row_num = 1
    for row_values in sheet.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if not any(row_values):
            continue
            
        ts_cell = row_values[header_indices['dấu thời gian'] - 1]
        op_name_cell = row_values[header_indices['họ tên người vận hành'] - 1]
        vcode_cell = row_values[header_indices['tên phương tiện'] - 1]
        hm_cell = row_values[header_indices['số giờ hoạt động'] - 1]
        
        if not ts_cell or not op_name_cell or not vcode_cell:
            if ts_cell or op_name_cell or vcode_cell:
                stats["errors"].append({
                    "row": row_num,
                    "message": "Thiếu thông tin bắt buộc (Dấu thời gian, Tên tài xế hoặc Xe), bỏ qua dòng này."
                })
            continue
            
        # Parse timestamp
        if isinstance(ts_cell, datetime):
            ts = ts_cell
        else:
            try:
                ts = datetime.strptime(str(ts_cell).strip(), "%d/%m/%Y %H:%M:%S")
            except ValueError:
                try:
                    ts = datetime.strptime(str(ts_cell).strip(), "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    stats["errors"].append({
                        "row": row_num,
                        "message": f"Thời gian '{ts_cell}' không đúng định dạng ngày giờ, bỏ qua."
                    })
                    continue
                    
        stats["rows_processed"] += 1
        
        # Get or create operator
        operator = get_or_create_operator(db, str(op_name_cell).strip(), "NGƯỜI VẬN HÀNH")
        
        vcode_raw = str(vcode_cell).strip()
        vcode_list = [v.strip() for v in re.split(r'[,;/]', vcode_raw) if v.strip()]
        
        for vcode in vcode_list:
            # Determine vehicle type from code or default
            type_name = "Cần cẩu"
            vcode_lower = vcode.lower()
            if vcode_lower.startswith("xn") or "xe nâng" in vcode_lower or "n nâng" in vcode_lower:
                type_name = "Xe nâng"
            elif vcode_lower.startswith("xc") or vcode_lower.startswith("sk") or vcode_lower.startswith("pc") or "xe cuốc" in vcode_lower:
                type_name = "Xe cuốc"
            elif vcode_lower.startswith("xd") or "xe đào" in vcode_lower:
                type_name = "Xe đào"
                
            vehicle = get_or_create_vehicle(db, vcode, type_name)
            
            # Parse hourmeter
            hm_val = 0.0
            try:
                hm_val = float(hm_cell) if hm_cell is not None else 0.0
            except (ValueError, TypeError):
                pass
            if hm_val > float(vehicle.current_hourmeter):
                vehicle.current_hourmeter = hm_val
                db.commit()
                
            # Determine Shift
            shift_id = get_shift_from_time(db, ts.time())
            work_date = ts.date()
            
            # Determine safety status
            is_safe = True
            if 'safety_confirmed' in special_columns:
                safety_val = row_values[special_columns['safety_confirmed'] - 1]
                if safety_val:
                    s_val_str = str(safety_val).strip().lower()
                    if "không" in s_val_str or "ngưng" in s_val_str:
                        is_safe = False
    
            # Determine failure notes
            failure_note_val = None
            if 'failure_note' in special_columns:
                failure_note_val = row_values[special_columns['failure_note'] - 1]
                
            has_failure = False
            failure_desc = ""
            if failure_note_val:
                fn_str = str(failure_note_val).strip()
                if fn_str and fn_str.lower() not in ("bình thường", "bình thương", "binh thường", "none", "không", "-", ""):
                    has_failure = True
                    failure_desc = fn_str
    
            # Determine checklist results
            condition = "ok"
            results_to_save = []
            
            for col_idx, chk_id in checklist_item_ids.items():
                cell_val = row_values[col_idx - 1]
                result_bool = True
                note_text = ""
                
                # Check if this item is safety confirmation or failure notes
                item_name = checklist_item_cols[col_idx].lower()
                if "đảm bảo an toàn" in item_name:
                    if cell_val:
                        val_str = str(cell_val).strip().lower()
                        if "không" in val_str or "ngưng" in val_str:
                            result_bool = False
                            condition = "broken"
                            note_text = str(cell_val).strip()
                    else:
                        result_bool = True
                elif "hư hỏng trong ca" in item_name or "ghi chú hư hỏng" in item_name:
                    if cell_val:
                        val_str = str(cell_val).strip().lower()
                        if val_str not in ("bình thường", "bình thương", "binh thường", "none", "không", "-", ""):
                            result_bool = False
                            condition = "broken"
                            note_text = str(cell_val).strip()
                    else:
                        result_bool = True
                else:
                    if cell_val is not None:
                        val_str = str(cell_val).strip().lower()
                        if val_str and val_str not in ("ok", "đạt", "yes", "có", "1", "true"):
                            result_bool = False
                            condition = "broken"
                            note_text = str(cell_val).strip()
                
                results_to_save.append({
                    "checklist_id": chk_id,
                    "result": result_bool,
                    "note": note_text
                })
    
            if has_failure:
                condition = "broken"
                
            # Save Operation Log
            op_log = db.query(OperationLog).filter_by(
                vehicle_id=vehicle.vehicle_id,
                work_date=work_date,
                shift_id=shift_id
            ).first()
            
            if not op_log:
                # Determine ca end time
                shift = db.query(Shift).filter_by(shift_id=shift_id).first()
                end_time_val = shift.end_time if shift else (ts + timedelta(hours=8)).time()
                
                op_log = OperationLog(
                    vehicle_id=vehicle.vehicle_id,
                    operator_id=operator.operator_id,
                    shift_id=shift_id,
                    work_date=work_date,
                    start_hour=ts.time().replace(microsecond=0),
                    end_hour=end_time_val.replace(microsecond=0),
                    hourmeter_start=hm_val,
                    hourmeter_end=hm_val,  # Automatically close the ca
                    condition_before_shift=condition,
                    is_safety_confirmed=is_safe,
                    idempotency_key=uuid.uuid4(),
                    notes=f"{failure_desc if has_failure else 'Imported from Google Forms checklist survey'} (Người vận hành gốc: {str(op_name_cell).strip()})" if operator.operator_id == "UNKNOWN" and op_name_cell else (failure_desc if has_failure else "Imported from Google Forms checklist survey")
                )
                db.add(op_log)
                db.commit()
                db.refresh(op_log)
                stats["operation_logs_created"] += 1
                
                # Save Checklist Results
                for r_save in results_to_save:
                    db_cr = ChecklistResult(
                        operation_id=op_log.operation_id,
                        checklist_id=r_save["checklist_id"],
                        result=r_save["result"],
                        note=r_save["note"]
                    )
                    db.add(db_cr)
                    stats["checklist_results_created"] += 1
                db.commit()
    
                # Create failure log if failure is reported
                if has_failure:
                    category_id = get_failure_category(db, failure_desc)
                    
                    # Check duplicate failure log
                    exist_fail = db.query(FailureLog).filter(
                        FailureLog.vehicle_id == vehicle.vehicle_id,
                        FailureLog.description == failure_desc,
                        func.date(FailureLog.failure_time) == work_date
                    ).first()
                    
                    if not exist_fail:
                        fail_log = FailureLog(
                            operation_id=op_log.operation_id,
                            vehicle_id=vehicle.vehicle_id,
                            category_id=category_id,
                            description=failure_desc,
                            failure_time=ts,
                            severity="dangerous" if not is_safe or any(k in failure_desc.lower() for k in ("nguy hiểm", "mất phanh", "hỏng phanh", "cháy", "gãy", "đứt cáp")) else "light",
                            phase="before_shift",
                            is_repaired=is_safe,
                            created_by=operator.operator_id
                        )
                        db.add(fail_log)
                        db.commit()
                        db.refresh(fail_log)
                        stats["failure_logs_created"] += 1
                        
                        # No Repair Log is created on import as per requirements
                        
                        if not is_safe:
                            vehicle.status = "repairing"
                            db.commit()

    final_vehicle_count = db.query(Vehicle).count()
    final_operator_count = db.query(Operator).count()
    final_checklist_count = db.query(ChecklistItem).count()
    
    stats["vehicles_created"] = final_vehicle_count - initial_vehicle_count
    stats["operators_created"] = final_operator_count - initial_operator_count
    stats["checklist_items_created"] = final_checklist_count - initial_checklist_count

    return {
        "message": "Nhập báo cáo checklist đầu ca thành công",
        "statistics": stats
    }

@router.post("/weekly-report")
async def import_weekly_report_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Operator = Depends(PermissionChecker(["reports:view"]))
):
    try:
        contents = await file.read()
        wrapper = ExcelWrapper(contents, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không thể mở tập tin Excel: {str(e)}")

    # Ensure required tables have statuses
    for code, label in {"ok": "Bình thường", "broken": "Có hư hỏng"}.items():
        if not db.query(ConditionStatus).filter_by(status_code=code).first():
            db.add(ConditionStatus(status_code=code, status_label=label))
    for code, label in {"pending": "Chờ sửa chữa", "in_progress": "Đang sửa chữa", "done": "Đã hoàn thành"}.items():
        if not db.query(RepairStatus).filter_by(status_code=code).first():
            db.add(RepairStatus(status_code=code, status_label=label))
            
    # Prefer shift with 'ngày' in its name, otherwise fallback to the first shift
    shift = db.query(Shift).filter(Shift.shift_name.ilike("%ngày%")).first()
    if not shift:
        shift = db.query(Shift).first()
    if not shift:
        shift = Shift(shift_id=1, shift_name="Ca Ngày", start_time=time(6, 0), end_time=time(18, 0))
        db.add(shift)
        db.commit()
        db.refresh(shift)
    default_shift_id = shift.shift_id

    stats = {
        "sheets_processed": 0,
        "sheets_skipped": [],
        "rows_processed": 0,
        "vehicles_created": 0,
        "operation_logs_created": 0,
        "failure_logs_created": 0,
        "repair_logs_created": 0,
        "errors": []
    }

    initial_vehicle_count = db.query(Vehicle).count()

    # Select only the target sheet (matching filename week, or the last valid sheet) to avoid importing historical sheets
    target_sheets = []
    # 1. Try to find week number in filename (e.g. T23, Tuần 23)
    week_match = re.search(r'(?:[Tt]uần|[Tt]|[Ww]eek|[Ww])[-_ ]*(\d+)', file.filename)
    if week_match:
        target_week = int(week_match.group(1))
        for name in wrapper.sheet_names:
            if any(kw in name.lower() for kw in ("huong dan", "hướng dẫn", "readme", "instruction", "guide")):
                continue
            sheet_week_match = re.search(r'(?:[Tt]uần|[Tt]|[Ww]eek|[Ww])?[-_ ]*(\d+)', name)
            if sheet_week_match:
                if int(sheet_week_match.group(1)) == target_week:
                    try:
                        sheet = wrapper.get_sheet(name)
                        header_row_idx, cols = find_header_row_and_cols(sheet, wrapper)
                        if header_row_idx and cols["name"]:
                            # Ensure the sheet has a valid date
                            work_date = extract_date_from_sheet(sheet, wrapper, sheet_name=name, filename=file.filename)
                            if work_date:
                                target_sheets = [name]
                                break
                    except Exception:
                        pass
                        
    # 2. Fallback: Find all valid weekly report sheets (must have valid date) and take the last one (newest week)
    if not target_sheets:
        valid_sheets = []
        for name in wrapper.sheet_names:
            if any(kw in name.lower() for kw in ("huong dan", "hướng dẫn", "readme", "instruction", "guide")):
                continue
            try:
                sheet = wrapper.get_sheet(name)
                header_row_idx, cols = find_header_row_and_cols(sheet, wrapper)
                if header_row_idx and cols["name"]:
                    work_date = extract_date_from_sheet(sheet, wrapper, sheet_name=name, filename=file.filename)
                    if work_date:
                        valid_sheets.append(name)
            except Exception:
                pass
        if valid_sheets:
            target_sheets = [valid_sheets[-1]]

    for sheet_name in target_sheets:

        sheet = wrapper.get_sheet(sheet_name)
        
        # 1. Parse date from sheet
        work_date = extract_date_from_sheet(sheet, wrapper, sheet_name=sheet_name, filename=file.filename)
        if not work_date:
            stats["errors"].append({
                "sheet": sheet_name,
                "row": 0,
                "message": "Không tìm thấy ngày báo cáo (dạng 'Ngày ... tháng ... năm ...') trong sheet, bỏ qua."
            })
            continue

        # Check if this week's date has already been imported from weekly report
        existing_log = db.query(OperationLog).filter(
            OperationLog.work_date == work_date,
            (OperationLog.notes.like("%Weekly Report%") | OperationLog.notes.like("%báo cáo tuần%"))
        ).first()
        if existing_log:
            stats["sheets_skipped"].append({
                "sheet": sheet_name,
                "date": str(work_date),
                "reason": "Tuần này đã được nhập dữ liệu trước đó."
            })
            continue

        # 2. Find header row and columns dynamically
        header_row_idx, cols = find_header_row_and_cols(sheet, wrapper)
        if not header_row_idx or not cols["name"]:
            # If the sheet doesn't look like a technical report sheet, skip it silently
            continue

        stats["sheets_processed"] += 1
        current_category = "Thiết bị"
        max_rows = wrapper.get_max_row(sheet)

        for r in range(header_row_idx + 1, max_rows + 1):
            stt_val = wrapper.get_cell_value(sheet, r, cols["stt"]) if cols["stt"] else None
            name_val = wrapper.get_cell_value(sheet, r, cols["name"]) if cols["name"] else None

            # Detect category separator rows
            row_cells_filled = 0
            for c_idx in range(1, wrapper.get_max_column(sheet) + 1):
                if wrapper.get_cell_value(sheet, r, c_idx) is not None:
                    row_cells_filled += 1

            if name_val and not stt_val and row_cells_filled <= 2:
                name_str = str(name_val).strip()
                if "cần cẩu" in name_str.lower() or "cẩu" in name_str.lower():
                    current_category = "Cần cẩu"
                elif "xe cuốc" in name_str.lower() or "cuốc" in name_str.lower():
                    current_category = "Xe cuốc"
                elif "xe nâng" in name_str.lower() or "nâng" in name_str.lower():
                    current_category = "Xe nâng"
                elif "gàu" in name_str.lower():
                    current_category = "Gàu hoa thị"
                else:
                    current_category = name_str
                continue

            if not name_val or str(name_val).strip() == "":
                continue

            vehicle_fullname = str(name_val).strip()

            # Guess vehicle code:
            vehicle_code = ""
            if cols["code"]:
                code_val = wrapper.get_cell_value(sheet, r, cols["code"])
                if code_val:
                    vehicle_code = str(code_val).strip()

            # Clean and fallback if code is empty
            if not vehicle_code:
                match = re.search(r'([A-Za-z0-9]+-[A-Za-z0-9]+|[A-Za-z0-9]+)$', vehicle_fullname)
                vehicle_code = match.group(1) if match else vehicle_fullname.split()[-1]

            vehicle_code = re.sub(r'\s+', '', vehicle_code)

            # Match vehicle in database (only active ones!)
            vehicle = db.query(Vehicle).filter(
                Vehicle.active == True,
                (
                    (func.replace(func.lower(Vehicle.vehicle_code), ' ', '') == vehicle_code.lower()) |
                    (func.lower(Vehicle.vehicle_name) == vehicle_fullname.lower())
                )
            ).first()

            if not vehicle:
                match = re.search(r'([A-Za-z0-9]+-[A-Za-z0-9]+|[A-Za-z0-9]+)$', vehicle_fullname)
                guessed_code = match.group(1) if match else vehicle_fullname.split()[-1]
                guessed_code_clean = re.sub(r'\s+', '', guessed_code)
                vehicle = db.query(Vehicle).filter(
                    Vehicle.active == True,
                    func.replace(func.lower(Vehicle.vehicle_code), ' ', '') == guessed_code_clean.lower()
                ).first()
                
            if not vehicle:
                # Alphanumeric fallback match
                code_alphanum = re.sub(r'[^A-Z0-9]', '', vehicle_code.upper())
                all_active_vehs = db.query(Vehicle).filter(Vehicle.active == True).all()
                for v in all_active_vehs:
                    v_alphanum = re.sub(r'[^A-Z0-9]', '', v.vehicle_code.upper())
                    if v_alphanum == code_alphanum:
                        vehicle = v
                        break

            if not vehicle:
                code_to_create = vehicle_code or vehicle_fullname
                vehicle = get_or_create_vehicle(db, code_to_create, current_category)
                vehicle.vehicle_name = vehicle_fullname
                db.commit()

            stats["rows_processed"] += 1

            # Parse Operating Hours
            hours_val = wrapper.get_cell_value(sheet, r, cols["hours"]) if cols["hours"] else None
            hours_float = 0.0
            if hours_val is not None:
                try:
                    hours_float = float(hours_val)
                except (ValueError, TypeError):
                    pass

            # Parse absolute Hourmeters if available
            hm_start = None
            hm_end = None
            if cols.get("hourmeter_start"):
                val = wrapper.get_cell_value(sheet, r, cols["hourmeter_start"])
                if val is not None:
                    try:
                        hm_start = float(val)
                    except (ValueError, TypeError):
                        pass
            if cols.get("hourmeter_end"):
                val = wrapper.get_cell_value(sheet, r, cols["hourmeter_end"])
                if val is not None:
                    try:
                        hm_end = float(val)
                    except (ValueError, TypeError):
                        pass

            if hm_start is not None and hm_end is not None:
                calculated_hours = hm_end - hm_start
                if hours_float == 0.0:
                    hours_float = max(0.0, calculated_hours)

            # Parse Active / Inactive checkmarks or status text
            is_active = True
            if cols["inactive"] and wrapper.get_cell_value(sheet, r, cols["inactive"]):
                val = str(wrapper.get_cell_value(sheet, r, cols["inactive"])).strip().lower()
                if val in ('x', '1', 'yes', 'có', 'true'):
                    is_active = False
            elif cols["active"] and wrapper.get_cell_value(sheet, r, cols["active"]):
                val = str(wrapper.get_cell_value(sheet, r, cols["active"])).strip().lower()
                if val in ('x', '1', 'yes', 'có', 'true'):
                    is_active = True
            elif cols["status_text"]:
                status_val = wrapper.get_cell_value(sheet, r, cols["status_text"])
                if status_val:
                    status_str = unicodedata.normalize("NFC", str(status_val).strip().lower())
                    if any(kw in status_str for kw in ("ngưng", "dừng", "hỏng", "dừng", "không hoạt động", "ngưng hoạt động", "chờ vật tư", "sửa chữa")):
                        is_active = False

            # Parse detailed columns
            desc_detail_val = wrapper.get_cell_value(sheet, r, cols["desc_detail"]) if cols["desc_detail"] else None
            desc_detail_str = str(desc_detail_val).strip() if desc_detail_val else ""

            downtime_desc_val = wrapper.get_cell_value(sheet, r, cols["downtime_desc"]) if cols["downtime_desc"] else None
            downtime_desc_str = str(downtime_desc_val).strip() if downtime_desc_val else ""
            if not downtime_desc_str and cols["downtime"]:
                downtime_val = wrapper.get_cell_value(sheet, r, cols["downtime"])
                if downtime_val is not None:
                    downtime_desc_str = str(downtime_val).strip()

            # 1. Parsing and correcting backlog_date (handling Excel regional date-swapped datetimes)
            backlog_date_val = wrapper.get_cell_value(sheet, r, cols["backlog_date"]) if cols["backlog_date"] else None
            backlog_date_str = ""
            if backlog_date_val:
                parsed_backlog_date = None
                if isinstance(backlog_date_val, (datetime, date)):
                    parsed_backlog_date = backlog_date_val if isinstance(backlog_date_val, date) else backlog_date_val.date()
                else:
                    parsed_backlog_date = parse_date(backlog_date_val)
                
                # Apply the swap heuristic: if parsed backlog_date is in the future relative to the report's work_date,
                # and day <= 12, swap day and month to get the correct past date.
                if parsed_backlog_date and work_date_str:
                    wdate = parse_date(work_date_str)
                    if wdate and parsed_backlog_date > wdate and parsed_backlog_date.day <= 12:
                        try:
                            parsed_backlog_date = date(parsed_backlog_date.year, parsed_backlog_date.day, parsed_backlog_date.month)
                        except ValueError:
                            pass
                
                if parsed_backlog_date:
                    backlog_date_str = parsed_backlog_date.strftime("%d/%m/%Y")

            suggestion_val = wrapper.get_cell_value(sheet, r, cols["suggestion"]) if cols["suggestion"] else None
            suggestion_str = str(suggestion_val).strip() if suggestion_val else ""

            notes_val = wrapper.get_cell_value(sheet, r, cols["notes"]) if cols["notes"] else None
            notes_str = str(notes_val).strip() if notes_val else ""
            
            # If downtime contains a text description, append it to notes_str
            if cols["downtime"]:
                dt_val = wrapper.get_cell_value(sheet, r, cols["downtime"])
                if dt_val is not None and not isinstance(dt_val, (int, float)):
                    dt_str = str(dt_val).strip()
                    dt_str_lower = dt_str.lower()
                    if re.search(r'\d{1,2}[/\-]\d{1,2}', dt_str_lower) or any(kw in dt_str_lower for kw in ("chưa", "không", "đang", "theo dõi", "chờ", "đứt", "báo từ", "tồn đọng")):
                        if notes_str:
                            notes_str = f"{notes_str} | TG dừng: {dt_str}"
                        else:
                            notes_str = f"TG dừng: {dt_str}"

            # Parse Condition & Create Failure / Repair logs
            cond_val = wrapper.get_cell_value(sheet, r, cols["condition"]) if cols["condition"] else None
            status_val = wrapper.get_cell_value(sheet, r, cols["status_text"]) if cols["status_text"] else None

            # Build failure description
            failure_desc_parts = []
            status_text_clean = ""
            db_status_text = ""
            if status_val and str(status_val).strip() and str(status_val).strip().lower() != "bình thường":
                status_text_clean = unicodedata.normalize("NFC", str(status_val).strip())
                status_text_clean_norm = status_text_clean.lower()
                
                # If it's "ngưng hẳn" or contains "ngưng hẳn", map to "Ngưng sửa chữa"
                if "ngưng hẳn" in status_text_clean_norm:
                    status_text_clean = re.sub(r'(?i)ngưng\s+hẳn', 'Ngưng sửa chữa', status_text_clean).strip()
                
                db_status_text = status_text_clean
                
                # Remove status phrases ("Ngưng sửa chữa", "Đang sửa chữa", "Chờ vật tư", "Có lỗi nhẹ", "Lỗi nhẹ", "Có sự cố") from description text
                status_text_clean = re.sub(r'(?i)ngưng\s+sửa\s+chữa|ngưng\s+sữa\s+chữa|đang\s+sửa\s+chữa|đang\s+sữa\s+chữa|đang\s+sc|chờ\s+vật\s+tư|chờ\s+vật\s+tư|chờ\s+vt|có\s+lỗi\s+nhẹ|có\s+lỗi\s+nhẹ|lỗi\s+nhẹ|lỗi\s+nhẹ|có\s+sự\s+cố|có\s+sự\s+cố', '', status_text_clean).strip()
                # Clean up leading/trailing symbols or spaces
                status_text_clean = re.sub(r'^[\s\-–—]+|[\s\-–—]+$', '', status_text_clean).strip()
                
                if status_text_clean:
                    failure_desc_parts.append(status_text_clean)
            if cond_val and str(cond_val).strip() and str(cond_val).strip().lower() not in [p.lower() for p in failure_desc_parts]:
                failure_desc_parts.append(str(cond_val).strip())

            # Fallback: if no condition description was found, use desc_detail_str as the main description
            used_desc_detail_in_meta = True
            if not failure_desc_parts and desc_detail_str:
                failure_desc_parts.append(desc_detail_str)
                used_desc_detail_in_meta = False

            main_desc = " - ".join(failure_desc_parts) if failure_desc_parts else "Bình thường"

            # Check if there is any failure/broken status
            is_broken = False
            if main_desc and main_desc.strip():
                norm_words = ("bình thường", "bình thương", "binh thường", "binh thuong", "ok", "hoạt động tốt", "hoạt động bình thường", "-", "")
                is_broken = main_desc.lower().strip() not in norm_words and not any(main_desc.lower().strip() == w for w in norm_words)

            # Build final detailed description string
            if is_broken:
                meta_parts = []
                if downtime_desc_str:
                    meta_parts.append(f"TG dừng: {downtime_desc_str}")
                if backlog_date_str:
                    meta_parts.append(f"Tồn đọng từ: {backlog_date_str}")
                if desc_detail_str and used_desc_detail_in_meta:
                    meta_parts.append(f"Chi tiết: {desc_detail_str}")
                if notes_str:
                    meta_parts.append(f"Ghi chú: {notes_str}")
                if suggestion_str:
                    meta_parts.append(f"Đề nghị: {suggestion_str}")

                if meta_parts:
                    cond_str = f"{main_desc} ({' | '.join(meta_parts)})"
                else:
                    cond_str = main_desc
            else:
                cond_str = main_desc

            # Check if failure was repaired during weekly sheet
            is_repaired = False
            if is_broken:
                if cols["repair_done"] and wrapper.get_cell_value(sheet, r, cols["repair_done"]):
                    done_val = str(wrapper.get_cell_value(sheet, r, cols["repair_done"])).strip().lower()
                    if any(ok_word in done_val for ok_word in ('x', '1', 'yes', 'có', 'đã xong', 'xong', 'rồi', 'done', 'true')):
                        if not any(neg in done_val for neg in ('chưa', 'không', 'đang')):
                            is_repaired = True
                    elif any(not_ok in done_val for not_ok in ('chưa', 'không', 'đang', 'chờ', 'theo dõi')):
                        is_repaired = False
                    else:
                        is_repaired = is_active
                else:
                    is_repaired = is_active

                # If the description has keywords indicating that the issue is still active or under monitoring,
                # mark it as not fully repaired (pending) unless it is explicitly marked complete.
                if any(kw in cond_str.lower() for kw in ("theo dõi", "chờ", "đang sửa", "chưa xong")):
                    explicit_done = False
                    if cols["repair_done"] and wrapper.get_cell_value(sheet, r, cols["repair_done"]):
                        done_val = str(wrapper.get_cell_value(sheet, r, cols["repair_done"])).strip().lower()
                        if any(ok_word in done_val for ok_word in ('đã xong', 'xong', 'rồi', 'done', 'có')) and not any(neg in done_val for neg in ('chưa', 'không')):
                            explicit_done = True
                    if not explicit_done:
                        is_repaired = False
            else:
                is_repaired = is_active

            # Update vehicle and active states based on whether the failure was resolved
            is_active_final = is_active
            if is_broken and not is_repaired:
                is_active_final = False

            if db_status_text.strip().lower() == "ngưng sửa chữa":
                vehicle.status = "stopped_repair"
            else:
                vehicle.status = "active" if is_active_final else "repairing"
            db.commit()

            # Ensure vehicle type/classification is updated if missing in database
            if not vehicle.vehicle_type_id:
                vtype = db.query(VehicleType).filter(func.lower(VehicleType.type_name) == current_category.lower()).first()
                if not vtype:
                    vtype = VehicleType(type_name=current_category)
                    db.add(vtype)
                    db.commit()
                    db.refresh(vtype)
                vehicle.vehicle_type_id = vtype.vehicle_type_id
                db.commit()

            # Parse downtime hours
            downtime_val = wrapper.get_cell_value(sheet, r, cols["downtime"]) if cols["downtime"] else None
            downtime_hours = 0.0
            if downtime_val is not None:
                if isinstance(downtime_val, (int, float)):
                    downtime_hours = float(downtime_val)
                else:
                    downtime_str = str(downtime_val).strip().lower()
                    # If it looks like a date/date-range (contains / or - between digits, e.g. 10/03)
                    if re.search(r'\d{1,2}[/\-]\d{1,2}', downtime_str):
                        downtime_hours = 0.0
                    else:
                        day_match = re.search(r'(\d+(?:\.\d+)?)\s*ngày', downtime_str)
                        if day_match:
                            downtime_hours = float(day_match.group(1)) * 24.0
                        else:
                            hour_match = re.search(r'(\d+(?:\.\d+)?)\s*(?:h|giờ|giờ)?', downtime_str)
                            if hour_match:
                                downtime_hours = float(hour_match.group(1))
                            else:
                                downtime_hours = 0.0

            # Notes already parsed at top of loop

            # Create or update OperationLog
            op_log = db.query(OperationLog).filter_by(
                vehicle_id=vehicle.vehicle_id,
                work_date=work_date
            ).first()

            # Determine start and end hourmeter values to save
            log_hm_start = Decimal(str(hm_start)) if hm_start is not None else vehicle.current_hourmeter
            log_hm_end = Decimal(str(hm_end)) if hm_end is not None else (log_hm_start + Decimal(str(hours_float)))

            # Use current logged in user as the operator who reports this
            operator = current_user

            # Formulate the notes string to always contain "[Weekly Report]" to prevent duplicates on re-import
            weekly_notes = f"[Weekly Report] {notes_str or cond_str}".strip() if (notes_str or cond_str) else "Imported from Weekly Report"

            if op_log:
                op_log.hourmeter_start = log_hm_start
                op_log.hourmeter_end = log_hm_end
                op_log.condition_before_shift = "ok" if (not is_broken or is_repaired) else "broken"
                op_log.is_safety_confirmed = is_active_final
                # Prepend [Weekly Report] if it's not already in there
                existing_notes = op_log.notes or ""
                if "Weekly Report" not in existing_notes and "báo cáo tuần" not in existing_notes:
                    op_log.notes = f"[Weekly Report] {existing_notes}".strip()
                if notes_str or cond_str:
                    op_log.notes = weekly_notes
                db.commit()

                if log_hm_end > vehicle.current_hourmeter:
                    vehicle.current_hourmeter = log_hm_end
                    db.commit()
            else:
                op_log = OperationLog(
                    vehicle_id=vehicle.vehicle_id,
                    operator_id=operator.operator_id,
                    shift_id=default_shift_id,
                    work_date=work_date,
                    start_hour=time(0, 0),
                    end_hour=time(23, 59),
                    hourmeter_start=log_hm_start,
                    hourmeter_end=log_hm_end,
                    condition_before_shift="ok" if (not is_broken or is_repaired) else "broken",
                    is_safety_confirmed=is_active_final,
                    idempotency_key=uuid.uuid4(),
                    notes=weekly_notes,
                    work_type="production"
                )
                db.add(op_log)
                db.commit()
                db.refresh(op_log)
                stats["operation_logs_created"] += 1

                if log_hm_end > vehicle.current_hourmeter:
                    vehicle.current_hourmeter = log_hm_end
                    db.commit()

            # Handle failures
            if is_broken:
                # Parse backlog date as failure_time date, fallback to work_date
                failure_date = work_date
                if backlog_date_val:
                    if isinstance(backlog_date_val, (datetime, date)):
                        failure_date = backlog_date_val.date() if isinstance(backlog_date_val, datetime) else backlog_date_val
                    else:
                        try:
                            parsed_backlog = parse_date(str(backlog_date_val).strip())
                            if parsed_backlog:
                                failure_date = parsed_backlog
                        except Exception:
                            pass

                start_dt = datetime.combine(failure_date, time.min)
                end_dt = datetime.combine(failure_date, time.max)
                exist_fail = db.query(FailureLog).filter(
                    FailureLog.vehicle_id == vehicle.vehicle_id,
                    FailureLog.description == cond_str,
                    FailureLog.failure_time >= start_dt,
                    FailureLog.failure_time <= end_dt
                ).first()

                is_repaired = False
                if cols["repair_done"] and wrapper.get_cell_value(sheet, r, cols["repair_done"]):
                    done_val = str(wrapper.get_cell_value(sheet, r, cols["repair_done"])).strip().lower()
                    if any(ok_word in done_val for ok_word in ('x', '1', 'yes', 'có', 'đã xong', 'xong', 'rồi', 'done', 'true')):
                        if not any(neg in done_val for neg in ('chưa', 'không', 'đang')):
                            is_repaired = True
                    elif any(not_ok in done_val for not_ok in ('chưa', 'không', 'đang', 'chờ', 'theo dõi')):
                        is_repaired = False
                    else:
                        is_repaired = is_active
                else:
                    is_repaired = is_active

                # If the description has keywords indicating that the issue is still active or under monitoring,
                # mark it as not fully repaired (pending) unless it is explicitly marked complete.
                if any(kw in cond_str.lower() for kw in ("theo dõi", "chờ", "đang sửa", "chưa xong")):
                    explicit_done = False
                    if cols["repair_done"] and wrapper.get_cell_value(sheet, r, cols["repair_done"]):
                        done_val = str(wrapper.get_cell_value(sheet, r, cols["repair_done"])).strip().lower()
                        if any(ok_word in done_val for ok_word in ('đã xong', 'xong', 'rồi', 'done', 'có')) and not any(neg in done_val for neg in ('chưa', 'không')):
                            explicit_done = True
                    if not explicit_done:
                        is_repaired = False

                if exist_fail:
                    exist_fail.is_repaired = is_repaired
                    db.commit()
                else:
                    cat_id = get_failure_category(db, cond_str)
                    f_time = datetime.combine(failure_date, time(12, 0))

                    new_fail = FailureLog(
                        operation_id=op_log.operation_id,
                        vehicle_id=vehicle.vehicle_id,
                        created_by=operator.operator_id,
                        category_id=cat_id,
                        description=cond_str,
                        failure_time=f_time,
                        is_repaired=is_repaired,
                        severity="dangerous" if not is_active or any(k in cond_str.lower() for k in ("nguy hiểm", "mất phanh", "hỏng phanh", "cháy", "gãy", "đứt cáp")) else "light",
                        phase="out_of_shift"
                    )
                    db.add(new_fail)
                    db.commit()
                    db.refresh(new_fail)
                    stats["failure_logs_created"] += 1

                    # No Repair Log is created on import as per requirements

    stats["vehicles_created"] = db.query(Vehicle).count() - initial_vehicle_count

    if stats["sheets_processed"] == 0:
        if len(stats["sheets_skipped"]) > 0:
            msg = "Tất cả các trang tính trong tệp đã được nhập dữ liệu trước đó. Vui lòng xóa lịch sử tuần này nếu muốn nhập lại."
        else:
            msg = "Không tìm thấy dữ liệu báo cáo tuần hợp lệ hoặc dòng tiêu đề thiết bị trong tệp Excel này."
    else:
        msg = "Nhập báo cáo tình trạng tuần thành công!"

    return {
        "message": msg,
        "statistics": stats
    }

@router.get("/weekly-report/history")
def get_weekly_report_import_history(
    db: Session = Depends(get_db),
    current_user: Operator = Depends(PermissionChecker(["reports:view"]))
):
    try:
        from sqlalchemy import case
        results = db.query(
            OperationLog.work_date,
            func.count(OperationLog.operation_id).label("logs_count"),
            func.sum(case((OperationLog.condition_before_shift == "broken", 1), else_=0)).label("failures_count")
        ).group_by(OperationLog.work_date).order_by(OperationLog.work_date.desc()).all()

        history = []
        for r in results:
            week_num = r.work_date.isocalendar()[1]
            history.append({
                "date": str(r.work_date),
                "week_name": f"Tuần {week_num}",
                "logs_count": r.logs_count,
                "failures_count": int(r.failures_count or 0)
            })
        return history
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi khi tải lịch sử import: {str(e)}")

@router.delete("/weekly-report/{work_date}")
def delete_weekly_report_import(
    work_date: date,
    db: Session = Depends(get_db),
    current_user: Operator = Depends(PermissionChecker(["reports:view"]))
):
    try:
        op_logs = db.query(OperationLog).filter(OperationLog.work_date == work_date).all()
        op_log_ids = [op.operation_id for op in op_logs]
        
        # Deduct hours from vehicle current_hourmeter when deleting
        for op in op_logs:
            if op.hourmeter_start is not None and op.hourmeter_end is not None:
                hours_run = float(op.hourmeter_end - op.hourmeter_start)
                if hours_run > 0:
                    vehicle = db.query(Vehicle).filter(Vehicle.vehicle_id == op.vehicle_id).first()
                    if vehicle:
                        vehicle.current_hourmeter = Decimal(str(max(0.0, float(vehicle.current_hourmeter) - hours_run)))
        db.commit()

        if op_log_ids:
            failures = db.query(FailureLog).filter(FailureLog.operation_id.in_(op_log_ids)).all()
            failure_ids = [f.failure_id for f in failures]
            
            if failure_ids:
                db.query(RepairLog).filter(RepairLog.failure_id.in_(failure_ids)).delete(synchronize_session=False)
                db.query(FailureLog).filter(FailureLog.failure_id.in_(failure_ids)).delete(synchronize_session=False)
                
        db.query(OperationLog).filter(OperationLog.work_date == work_date).delete(synchronize_session=False)
        db.commit()

        return {"message": f"Xóa dữ liệu tuần {work_date} thành công."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi khi xóa dữ liệu tuần: {str(e)}")
