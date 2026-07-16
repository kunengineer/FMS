from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from app.core.database import get_db
from app.core.dependencies import PermissionChecker, get_current_user
from app.models import Vehicle, OperationLog, FailureLog, FailureCategory, RepairLog, Operator, ChecklistItem, ChecklistResult
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from typing import Optional
from datetime import datetime, date, timedelta
from decimal import Decimal

router = APIRouter(prefix="/reports", tags=["reports"])

@router.get("/metrics")
def get_reports_metrics(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: Operator = Depends(PermissionChecker(["reports:view"]))
):
    """
    Get raw reports metrics (useful for viewing tables in UI before exporting to Excel)
    """
    # 1. MTTR and MTBF calculations per vehicle
    vehicles = db.query(Vehicle).filter(Vehicle.active == True).all()
    mttr_mtbf_list = []
    
    for v in vehicles:
        # MTTR = AVG(repair_end - repair_start) in hours for done repairs
        repairs = db.query(RepairLog).join(
            FailureLog, FailureLog.failure_id == RepairLog.failure_id
        ).filter(
            FailureLog.vehicle_id == v.vehicle_id,
            RepairLog.repair_status == "done",
            RepairLog.repair_end != None
        ).all()
        
        if repairs:
            total_repair_hours = sum((r.repair_end - r.repair_start).total_seconds() / 3600 for r in repairs)
            mttr = round(total_repair_hours / len(repairs), 1)
        else:
            mttr = 0.0
            
        # MTBF = AVG(interval between consecutive failures) for resolved failures
        failures = db.query(FailureLog).filter(
            FailureLog.vehicle_id == v.vehicle_id,
            FailureLog.is_repaired == True
        ).order_by(FailureLog.failure_time.asc()).all()
        
        if len(failures) >= 2:
            intervals = []
            for i in range(len(failures) - 1):
                delta = (failures[i+1].failure_time - failures[i].failure_time).total_seconds() / 3600
                intervals.append(delta)
            mtbf = round(sum(intervals) / len(intervals), 1)
        else:
            mtbf = "N/A"

        # KPI active hours, repair count, downtime
        op_logs_query = db.query(OperationLog).filter(OperationLog.vehicle_id == v.vehicle_id)
        if start_date:
            op_logs_query = op_logs_query.filter(OperationLog.work_date >= start_date)
        if end_date:
            op_logs_query = op_logs_query.filter(OperationLog.work_date <= end_date)
        op_logs = op_logs_query.all()
        
        total_hours = sum((op.hourmeter_end - op.hourmeter_start) for op in op_logs if op.hourmeter_end is not None)
        active_hours_val = float(total_hours)

        r_query = db.query(RepairLog).join(FailureLog, FailureLog.failure_id == RepairLog.failure_id)\
                                     .filter(FailureLog.vehicle_id == v.vehicle_id, RepairLog.repair_status == "done")
        if start_date:
            r_query = r_query.filter(RepairLog.repair_end >= datetime.combine(start_date, datetime.min.time()))
        if end_date:
            r_query = r_query.filter(RepairLog.repair_end <= datetime.combine(end_date, datetime.max.time()))
        repair_count = r_query.count()

        downtime_hours = get_vehicle_downtime_hours(db, v.vehicle_id, start_date, end_date)
        
        if active_hours_val > 0:
            efficiency_rate = round((downtime_hours / active_hours_val) * 100, 1)
        else:
            efficiency_rate = 0.0
            
        mttr_mtbf_list.append({
            "vehicle_code": v.vehicle_code,
            "vehicle_name": v.vehicle_name,
            "mttr": mttr,
            "mtbf": mtbf,
            "active_hours": active_hours_val,
            "repair_count": repair_count,
            "downtime_hours": downtime_hours,
            "efficiency_rate": efficiency_rate
        })
        
    return {
        "mttr_mtbf": mttr_mtbf_list
    }

def get_vehicle_downtime_hours(db: Session, vehicle_id, start_date, end_date) -> float:
    import uuid
    f_query = db.query(FailureLog).filter(FailureLog.vehicle_id == vehicle_id)
    if start_date:
        f_query = f_query.filter(FailureLog.failure_time >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        f_query = f_query.filter(FailureLog.failure_time <= datetime.combine(end_date, datetime.max.time()))
    failures = f_query.all()
    
    total_downtime = 0.0
    now_dt = datetime.now()
    if end_date:
        now_dt = datetime.combine(end_date, datetime.max.time())
        
    for f in failures:
        if f.is_repaired:
            repair = db.query(RepairLog).filter(
                RepairLog.failure_id == f.failure_id,
                RepairLog.repair_status == "done",
                RepairLog.repair_end != None
            ).first()
            if repair:
                dt = (repair.repair_end - repair.repair_start).total_seconds() / 3600.0
                total_downtime += max(dt, 0.0)
        else:
            repair = db.query(RepairLog).filter(
                RepairLog.failure_id == f.failure_id,
                RepairLog.repair_status == "in_progress"
            ).first()
            start_dt = repair.repair_start if repair else f.failure_time
            if start_dt < now_dt:
                dt = (now_dt - start_dt).total_seconds() / 3600.0
                total_downtime += max(dt, 0.0)
                
    return round(total_downtime, 1)

@router.get("/export")
def export_reports_excel(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: Operator = Depends(PermissionChecker(["reports:view"]))
):
    # Create Excel Workbook
    wb = openpyxl.Workbook()
    
    # Stylings
    header_fill = PatternFill(start_color="1A56DB", end_color="1A56DB", fill_type="solid") # Deep Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=16, bold=True, color="1A56DB")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    def style_sheet(ws, title_text, headers):
        ws.views.sheetView[0].showGridLines = True
        
        # Add Title Row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.row_dimensions[1].height = 40
        cell = ws.cell(row=1, column=1)
        cell.value = title_text
        cell.font = title_font
        cell.alignment = align_left
        
        # Add headers
        ws.row_dimensions[3].height = 25
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col_idx)
            c.value = h
            c.fill = header_fill
            c.font = header_font
            c.alignment = align_center
            c.border = border_thin

    # --- SHEET 1: HƯ HỎNG THEO HẠNG MỤC ---
    ws1 = wb.active
    ws1.title = "Hư Hỏng Theo Hạng Mục"
    
    # Query failures by category
    query_s1 = db.query(
        FailureCategory.category_name,
        func.count(FailureLog.failure_id).label("cnt")
    ).join(FailureLog, FailureLog.category_id == FailureCategory.category_id)
    if start_date:
        query_s1 = query_s1.filter(FailureLog.failure_time >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query_s1 = query_s1.filter(FailureLog.failure_time <= datetime.combine(end_date, datetime.max.time()))
    s1_results = query_s1.group_by(FailureCategory.category_name).all()
    
    total_failures = sum(r[1] for r in s1_results) or 1
    
    headers1 = ["STT", "Hạng Mục Hư Hỏng", "Số Lượng Sự Cố", "Tỷ Lệ (%)"]
    style_sheet(ws1, "BÁO CÁO HƯ HỎNG THEO HẠNG MỤC", headers1)
    
    row_idx = 4
    for idx, (cat_name, count) in enumerate(s1_results, 1):
        ws1.row_dimensions[row_idx].height = 20
        # values
        ws1.cell(row=row_idx, column=1, value=idx).alignment = align_center
        ws1.cell(row=row_idx, column=2, value=cat_name).alignment = align_left
        ws1.cell(row=row_idx, column=3, value=count).alignment = align_right
        ws1.cell(row=row_idx, column=4, value=round((count / total_failures) * 100, 1)).alignment = align_right
        
        # borders and fonts
        for c in range(1, 5):
            cell = ws1.cell(row=row_idx, column=c)
            cell.font = data_font
            cell.border = border_thin
        row_idx += 1

    # --- SHEET 2: TOP XE HƯ NHIỀU NHẤT ---
    ws2 = wb.create_sheet("Top Xe Hư Nhiều Nhất")
    query_s2 = db.query(
        Vehicle.vehicle_code,
        Vehicle.vehicle_name,
        func.count(FailureLog.failure_id).label("cnt")
    ).join(FailureLog, FailureLog.vehicle_id == Vehicle.vehicle_id).filter(Vehicle.active == True)
    if start_date:
        query_s2 = query_s2.filter(FailureLog.failure_time >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query_s2 = query_s2.filter(FailureLog.failure_time <= datetime.combine(end_date, datetime.max.time()))
    s2_results = query_s2.group_by(Vehicle.vehicle_code, Vehicle.vehicle_name).order_by(func.count(FailureLog.failure_id).desc()).all()
    
    headers2 = ["STT", "Mã Phương Tiện", "Tên Phương Tiện", "Số Lần Gặp Sự Cố"]
    style_sheet(ws2, "DANH SÁCH PHƯƠNG TIỆN CÓ SỰ CỐ NHIỀU NHẤT", headers2)
    
    row_idx = 4
    for idx, (code, name, count) in enumerate(s2_results, 1):
        ws2.row_dimensions[row_idx].height = 20
        ws2.cell(row=row_idx, column=1, value=idx).alignment = align_center
        ws2.cell(row=row_idx, column=2, value=code).alignment = align_center
        ws2.cell(row=row_idx, column=3, value=name).alignment = align_left
        ws2.cell(row=row_idx, column=4, value=count).alignment = align_right
        
        for c in range(1, 5):
            cell = ws2.cell(row=row_idx, column=c)
            cell.font = data_font
            cell.border = border_thin
        row_idx += 1

    # --- SHEET 3: THỜI GIAN DỪNG MÁY (MTTR, MTBF) ---
    ws3 = wb.create_sheet("Chỉ số MTTR - MTBF")
    headers3 = ["STT", "Mã Phương Tiện", "Tên Phương Tiện", "MTTR (Giờ Sửa Chữa TB)", "MTBF (Thời Gian Giữa 2 Sự Cố TB)"]
    style_sheet(ws3, "THỜI GIAN DỪNG MÁY & CHI TIẾT HIỆU SUẤT VẬN HÀNH (MTTR/MTBF)", headers3)
    
    vehicles = db.query(Vehicle).filter(Vehicle.active == True).all()
    row_idx = 4
    for idx, v in enumerate(vehicles, 1):
        # MTTR
        repairs = db.query(RepairLog).join(
            FailureLog, FailureLog.failure_id == RepairLog.failure_id
        ).filter(
            FailureLog.vehicle_id == v.vehicle_id,
            RepairLog.repair_status == "done",
            RepairLog.repair_end != None
        )
        if start_date:
            repairs = repairs.filter(RepairLog.repair_start >= datetime.combine(start_date, datetime.min.time()))
        if end_date:
            repairs = repairs.filter(RepairLog.repair_end <= datetime.combine(end_date, datetime.max.time()))
        repairs = repairs.all()
        
        if repairs:
            total_repair_hours = sum((r.repair_end - r.repair_start).total_seconds() / 3600 for r in repairs)
            mttr_val = round(total_repair_hours / len(repairs), 1)
        else:
            mttr_val = 0.0
            
        # MTBF
        failures = db.query(FailureLog).filter(
            FailureLog.vehicle_id == v.vehicle_id,
            FailureLog.is_repaired == True
        )
        if start_date:
            failures = failures.filter(FailureLog.failure_time >= datetime.combine(start_date, datetime.min.time()))
        if end_date:
            failures = failures.filter(FailureLog.failure_time <= datetime.combine(end_date, datetime.max.time()))
        failures = failures.order_by(FailureLog.failure_time.asc()).all()
        
        if len(failures) >= 2:
            intervals = []
            for i in range(len(failures) - 1):
                delta = (failures[i+1].failure_time - failures[i].failure_time).total_seconds() / 3600
                intervals.append(delta)
            mtbf_val = round(sum(intervals) / len(intervals), 1)
        else:
            mtbf_val = "N/A"
            
        ws3.row_dimensions[row_idx].height = 20
        ws3.cell(row=row_idx, column=1, value=idx).alignment = align_center
        ws3.cell(row=row_idx, column=2, value=v.vehicle_code).alignment = align_center
        ws3.cell(row=row_idx, column=3, value=v.vehicle_name).alignment = align_left
        ws3.cell(row=row_idx, column=4, value=mttr_val).alignment = align_right
        ws3.cell(row=row_idx, column=5, value=mtbf_val).alignment = align_right
        
        for c in range(1, 6):
            cell = ws3.cell(row=row_idx, column=c)
            cell.font = data_font
            cell.border = border_thin
        row_idx += 1

    # --- SHEET 4: TẦN SUẤT HƯ HỎNG LẶP LẠI ---
    ws4 = wb.create_sheet("Hư Hỏng Lặp Lại")
    headers4 = ["STT", "Mã Phương Tiện", "Tên Phương Tiện", "Hạng Mục Hư Hỏng", "Số Lần Xuất Hiện"]
    style_sheet(ws4, "TẦN SUẤT HƯ HỎNG LẶP LẠI THEO PHƯƠNG TIỆN", headers4)
    
    query_s4 = db.query(
        Vehicle.vehicle_code,
        Vehicle.vehicle_name,
        FailureCategory.category_name,
        func.count(FailureLog.failure_id).label("cnt")
    ).join(FailureLog, FailureLog.vehicle_id == Vehicle.vehicle_id)\
     .join(FailureCategory, FailureCategory.category_id == FailureLog.category_id)\
     .filter(Vehicle.active == True)
    if start_date:
        query_s4 = query_s4.filter(FailureLog.failure_time >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query_s4 = query_s4.filter(FailureLog.failure_time <= datetime.combine(end_date, datetime.max.time()))
    s4_results = query_s4.group_by(Vehicle.vehicle_code, Vehicle.vehicle_name, FailureCategory.category_name)\
                         .order_by(func.count(FailureLog.failure_id).desc()).all()
                         
    row_idx = 4
    for idx, (v_code, v_name, cat_name, count) in enumerate(s4_results, 1):
        ws4.row_dimensions[row_idx].height = 20
        ws4.cell(row=row_idx, column=1, value=idx).alignment = align_center
        ws4.cell(row=row_idx, column=2, value=v_code).alignment = align_center
        ws4.cell(row=row_idx, column=3, value=v_name).alignment = align_left
        ws4.cell(row=row_idx, column=4, value=cat_name).alignment = align_left
        ws4.cell(row=row_idx, column=5, value=count).alignment = align_right
        
        for c in range(1, 6):
            cell = ws4.cell(row=row_idx, column=c)
            cell.font = data_font
            cell.border = border_thin
        row_idx += 1

    # --- SHEET 5: TỔNG HỢP KPI ---
    ws5 = wb.create_sheet("Tổng Hợp KPI Đội Xe")
    headers5 = [
        "STT", "Mã Phương Tiện", "Tên Phương Tiện", "Số Ca Hoạt Động", 
        "Tổng Giờ Máy Chạy", "Số Lần Sự Cố", "Số Lần Sửa Chữa", "Tỷ Lệ Sự Cố/Ca (%)",
        "Tổng Giờ Hư Hỏng (Downtime) (Giờ)", "Hiệu Suất Vận Hành (Hư/Hoạt Động) (%)"
    ]
    style_sheet(ws5, "TỔNG HỢP KPI HIỆU SUẤT HOẠT ĐỘNG CỦA ĐỘI XE", headers5)
    
    row_idx = 4
    for idx, v in enumerate(vehicles, 1):
        # Shifts count
        op_logs_query = db.query(OperationLog).filter(OperationLog.vehicle_id == v.vehicle_id)
        if start_date:
            op_logs_query = op_logs_query.filter(OperationLog.work_date >= start_date)
        if end_date:
            op_logs_query = op_logs_query.filter(OperationLog.work_date <= end_date)
        op_logs = op_logs_query.all()
        
        shift_count = len(op_logs)
        
        # Total hours = SUM(hourmeter_end - hourmeter_start)
        total_hours = sum((op.hourmeter_end - op.hourmeter_start) for op in op_logs if op.hourmeter_end is not None)
        
        # Failure count
        f_query = db.query(FailureLog).filter(FailureLog.vehicle_id == v.vehicle_id)
        if start_date:
            f_query = f_query.filter(FailureLog.failure_time >= datetime.combine(start_date, datetime.min.time()))
        if end_date:
            f_query = f_query.filter(FailureLog.failure_time <= datetime.combine(end_date, datetime.max.time()))
        failure_count = f_query.count()
        
        # Repair count
        r_query = db.query(RepairLog).join(FailureLog, FailureLog.failure_id == RepairLog.failure_id)\
                                     .filter(FailureLog.vehicle_id == v.vehicle_id, RepairLog.repair_status == "done")
        if start_date:
            r_query = r_query.filter(RepairLog.repair_end >= datetime.combine(start_date, datetime.min.time()))
        if end_date:
            r_query = r_query.filter(RepairLog.repair_end <= datetime.combine(end_date, datetime.max.time()))
        repair_count = r_query.count()
        
        failure_rate = round((failure_count / shift_count) * 100, 1) if shift_count > 0 else 0.0
        
        # Calculate downtime and efficiency rate
        downtime_hours = get_vehicle_downtime_hours(db, v.vehicle_id, start_date, end_date)
        active_hours_val = float(total_hours)
        if active_hours_val > 0:
            efficiency_rate = round((downtime_hours / active_hours_val) * 100, 1)
        else:
            efficiency_rate = 0.0
            
        ws5.row_dimensions[row_idx].height = 20
        ws5.cell(row=row_idx, column=1, value=idx).alignment = align_center
        ws5.cell(row=row_idx, column=2, value=v.vehicle_code).alignment = align_center
        ws5.cell(row=row_idx, column=3, value=v.vehicle_name).alignment = align_left
        ws5.cell(row=row_idx, column=4, value=shift_count).alignment = align_right
        ws5.cell(row=row_idx, column=5, value=active_hours_val).alignment = align_right
        ws5.cell(row=row_idx, column=6, value=failure_count).alignment = align_right
        ws5.cell(row=row_idx, column=7, value=repair_count).alignment = align_right
        ws5.cell(row=row_idx, column=8, value=failure_rate).alignment = align_right
        ws5.cell(row=row_idx, column=9, value=downtime_hours).alignment = align_right
        ws5.cell(row=row_idx, column=10, value=efficiency_rate).alignment = align_right
        
        for c in range(1, 11):
            cell = ws5.cell(row=row_idx, column=c)
            cell.font = data_font
            cell.border = border_thin
        row_idx += 1

    # --- SHEET 6: NHẬT KÝ CHECKLIST CHI TIẾT ---
    ws6 = wb.create_sheet("Nhật Ký Checklist Chi Tiết")
    
    # 1. Get all active checklist items (excluding special safety/failure columns)
    checklist_items = db.query(ChecklistItem).filter(
        ChecklistItem.active == True,
        ~ChecklistItem.item_name.like("%đảm bảo an toàn%"),
        ~ChecklistItem.item_name.like("%hư hỏng trong ca%"),
        ~ChecklistItem.item_name.like("%ghi chú hư hỏng%")
    ).order_by(ChecklistItem.checklist_id.asc()).all()
    
    # 2. Build headers
    headers6 = [
        "Dấu thời gian", "Điểm số (Lỗi/Tổng)", "Họ Tên Người Vận Hành", 
        "Mã Thiết Bị", "Tên Thiết Bị", "Số giờ máy đầu ca"
    ]
    for item in checklist_items:
        headers6.append(item.item_name)
    headers6.extend(["Sự cố ghi nhận trong ca", "Đảm bảo an toàn bắt đầu ca"])
    
    style_sheet(ws6, "NHẬT KÝ CHI TIẾT KẾT QUẢ KIỂM TRA CHECKLIST TRƯỚC CA CHẠY XE", headers6)
    
    # 3. Fetch all operation logs
    op_logs_query = db.query(OperationLog).options(
        joinedload(OperationLog.operator), 
        joinedload(OperationLog.vehicle),
        joinedload(OperationLog.checklist_results).joinedload(ChecklistResult.checklist_item),
        joinedload(OperationLog.failures)
    ).order_by(OperationLog.work_date.asc(), OperationLog.start_hour.asc())
    
    if start_date:
        op_logs_query = op_logs_query.filter(OperationLog.work_date >= start_date)
    if end_date:
        op_logs_query = op_logs_query.filter(OperationLog.work_date <= end_date)
        
    op_logs = op_logs_query.all()
    
    row_idx = 4
    for idx, log in enumerate(op_logs, 1):
        ws6.row_dimensions[row_idx].height = 20
        
        # Datetime
        dt_str = f"{log.work_date.strftime('%d/%m/%Y')} {log.start_hour.strftime('%H:%M:%S')}"
        
        # Checklist counts
        total_checks = len(log.checklist_results)
        failed_checks = sum(1 for r in log.checklist_results if not r.result)
        score_str = f"{failed_checks} / {total_checks}"
        
        # Operators / Vehicles
        op_name = log.operator.full_name if log.operator else "N/A"
        v_code = log.vehicle.vehicle_code if log.vehicle else "N/A"
        v_name = log.vehicle.vehicle_name if log.vehicle else "N/A"
        hm_start = float(log.hourmeter_start)
        
        # Write basic columns
        ws6.cell(row=row_idx, column=1, value=dt_str).alignment = align_center
        ws6.cell(row=row_idx, column=2, value=score_str).alignment = align_center
        ws6.cell(row=row_idx, column=3, value=op_name).alignment = align_left
        ws6.cell(row=row_idx, column=4, value=v_code).alignment = align_center
        ws6.cell(row=row_idx, column=5, value=v_name).alignment = align_left
        ws6.cell(row=row_idx, column=6, value=hm_start).alignment = align_right
        
        # Results map
        results_map = {r.checklist_id: r for r in log.checklist_results}
        
        # Write checklist item columns
        col_offset = 7
        for item in checklist_items:
            res_obj = results_map.get(item.checklist_id)
            if res_obj is not None:
                val = "Đạt" if res_obj.result else "Không Đạt"
            else:
                val = "-"
            
            cell = ws6.cell(row=row_idx, column=col_offset, value=val)
            cell.alignment = align_center
            if val == "Không Đạt":
                cell.font = Font(name="Calibri", size=11, color="991B1B", bold=True)
            col_offset += 1
            
        # Failure notes
        failures_desc = "Bình thường"
        if log.failures:
            failures_desc = ", ".join(f.description for f in log.failures)
        else:
            failed_checks = []
            for item in checklist_items:
                res_obj = results_map.get(item.checklist_id)
                if res_obj is not None and not res_obj.result:
                    desc = f"{item.item_name}: {res_obj.note}" if res_obj.note else f"{item.item_name} (Không Đạt)"
                    failed_checks.append(desc)
            if failed_checks:
                failures_desc = "Lỗi checklist: " + ", ".join(failed_checks)
            elif log.notes and log.notes != "Imported from Google Forms checklist survey":
                failures_desc = log.notes
        ws6.cell(row=row_idx, column=col_offset, value=failures_desc).alignment = align_left
        
        # Safety confirmed status
        safety_status = "Đảm bảo" if log.is_safety_confirmed else "Không an toàn - yêu cầu dừng kiểm tra"
        safety_cell = ws6.cell(row=row_idx, column=col_offset + 1, value=safety_status)
        safety_cell.alignment = align_center
        if not log.is_safety_confirmed:
            safety_cell.font = Font(name="Calibri", size=11, color="991B1B", bold=True)
            
        # Set border and general font for basic/notes/safety columns
        total_cols = len(headers6)
        for c in range(1, total_cols + 1):
            cell = ws6.cell(row=row_idx, column=c)
            if cell.font.name != "Calibri" or cell.font.color is None:
                cell.font = data_font
            cell.border = border_thin
            
        row_idx += 1

    # Auto-adjust column width for all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                # avoid using title row for auto-width estimation
                if cell.row == 1:
                    continue
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to buffer and return as streaming response
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"bao_cao_hoat_dong_kpi_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/analytics")
def get_reports_analytics(
    timeframe: str = Query("week"),  # "day", "week", "quarter", "year"
    db: Session = Depends(get_db),
    current_user: Operator = Depends(PermissionChecker(["reports:view"]))
):
    """
    Get dynamic failure and repair analytics for various timeframes
    """
    # 1. Realtime vehicle production status (ready vs repairing vs inactive)
    total_vehicles = db.query(Vehicle).filter(Vehicle.active == True).count()
    active_vehicles = db.query(Vehicle).filter(Vehicle.active == True, Vehicle.status == "active").all()
    repairing_vehicles = db.query(Vehicle).filter(Vehicle.active == True, Vehicle.status == "repairing").all()
    stopped_repair_vehicles = db.query(Vehicle).filter(Vehicle.active == True, Vehicle.status == "stopped_repair").all()
    inactive_vehicles = db.query(Vehicle).filter(Vehicle.active == True, Vehicle.status == "inactive").all()

    # Create lists of codes/names for easy display
    status_summary = {
        "active_count": len(active_vehicles),
        "repairing_count": len(repairing_vehicles),
        "stopped_repair_count": len(stopped_repair_vehicles),
        "inactive_count": len(inactive_vehicles),
        "total_count": total_vehicles,
        "active_list": [{
            "vehicle_id": str(v.vehicle_id),
            "vehicle_code": v.vehicle_code,
            "vehicle_name": v.vehicle_name,
            "current_hourmeter": float(v.current_hourmeter),
            "status": "active",
            "status_text": "Có sự cố (Đã sửa)" if db.query(FailureLog).filter(FailureLog.vehicle_id == v.vehicle_id).first() is not None else "Đạt an toàn",
            "model": v.model,
            "manufacture_year": v.manufacture_year,
            "last_maintenance_hourmeter": float(v.last_maintenance_hourmeter) if v.last_maintenance_hourmeter else 0.0,
            "vehicle_type": {"type_name": v.vehicle_type.type_name} if v.vehicle_type else None
        } for v in active_vehicles],
        "repairing_list": [{
            "vehicle_id": str(v.vehicle_id),
            "vehicle_code": v.vehicle_code,
            "vehicle_name": v.vehicle_name,
            "current_hourmeter": float(v.current_hourmeter),
            "status": "repairing",
            "status_text": "Có sự cố (Đang sửa)",
            "model": v.model,
            "manufacture_year": v.manufacture_year,
            "last_maintenance_hourmeter": float(v.last_maintenance_hourmeter) if v.last_maintenance_hourmeter else 0.0,
            "vehicle_type": {"type_name": v.vehicle_type.type_name} if v.vehicle_type else None
        } for v in repairing_vehicles],
        "stopped_repair_list": [{
            "vehicle_id": str(v.vehicle_id),
            "vehicle_code": v.vehicle_code,
            "vehicle_name": v.vehicle_name,
            "current_hourmeter": float(v.current_hourmeter),
            "status": "stopped_repair",
            "status_text": "Có sự cố (Ngưng sửa chữa)",
            "model": v.model,
            "manufacture_year": v.manufacture_year,
            "last_maintenance_hourmeter": float(v.last_maintenance_hourmeter) if v.last_maintenance_hourmeter else 0.0,
            "vehicle_type": {"type_name": v.vehicle_type.type_name} if v.vehicle_type else None
        } for v in stopped_repair_vehicles],
        "inactive_list": [{
            "vehicle_id": str(v.vehicle_id),
            "vehicle_code": v.vehicle_code,
            "vehicle_name": v.vehicle_name,
            "current_hourmeter": float(v.current_hourmeter),
            "status": "inactive",
            "status_text": "Không đạt an toàn",
            "model": v.model,
            "manufacture_year": v.manufacture_year,
            "last_maintenance_hourmeter": float(v.last_maintenance_hourmeter) if v.last_maintenance_hourmeter else 0.0,
            "vehicle_type": {"type_name": v.vehicle_type.type_name} if v.vehicle_type else None
        } for v in inactive_vehicles]
    }

    # 2. Timeframe duration computation
    latest_failure = db.query(func.max(FailureLog.failure_time)).scalar()
    latest_op_log = db.query(func.max(OperationLog.work_date)).scalar()
    
    ref_time = datetime.now()
    
    db_latest = None
    if latest_failure:
        db_latest = latest_failure
    if latest_op_log:
        latest_op_dt = datetime.combine(latest_op_log, datetime.max.time())
        if not db_latest or latest_op_dt > db_latest:
            db_latest = latest_op_dt
            
    if db_latest and (datetime.now() - db_latest).days > 7:
        ref_time = db_latest

    if timeframe == "day":
        start_time = ref_time - timedelta(days=1)
        group_format = "%H:00" # group by hour
    elif timeframe == "week":
        start_time = ref_time - timedelta(weeks=1)
        group_format = "%d/%m" # group by day
    elif timeframe == "quarter":
        start_time = ref_time - timedelta(days=90)
        group_format = "%W/%y" # group by week
    elif timeframe == "year":
        start_time = ref_time - timedelta(days=365)
        group_format = "%m/%Y" # group by month
    else: # default to week
        start_time = ref_time - timedelta(weeks=1)
        group_format = "%d/%m"

    # 3. Total repaired vehicles and category counts in timeframe
    failures = db.query(FailureLog).options(joinedload(FailureLog.vehicle), joinedload(FailureLog.category)).filter(
        FailureLog.failure_time >= start_time
    ).all()

    vehicle_counts = {}
    category_counts = {}
    total_repairs = 0
    total_unresolved = 0

    for f in failures:
        vcode = f.vehicle.vehicle_code if f.vehicle else "Chưa rõ"
        vname = f.vehicle.vehicle_name if f.vehicle else "Chưa rõ"
        vehicle_counts[(vcode, vname)] = vehicle_counts.get((vcode, vname), 0) + 1

        catname = f.category.category_name if f.category else "Chưa rõ"
        category_counts[catname] = category_counts.get(catname, 0) + 1

        if f.is_repaired:
            total_repairs += 1
        else:
            total_unresolved += 1

    vehicle_freq = [
        {"vehicle_code": k[0], "vehicle_name": k[1], "count": v}
        for k, v in sorted(vehicle_counts.items(), key=lambda x: x[1], reverse=True)
    ]

    category_freq = [
        {"category_name": k, "count": v}
        for k, v in sorted(category_counts.items(), key=lambda x: x[1], reverse=True)
    ]

    timeline_map = {}
    for f in failures:
        label = f.failure_time.strftime(group_format)
        timeline_map[label] = timeline_map.get(label, 0) + 1

    timeline_data = []
    if timeframe == "day":
        for i in range(24):
            t = ref_time - timedelta(hours=i)
            lbl = t.strftime(group_format)
            timeline_data.insert(0, {"label": lbl, "count": timeline_map.get(lbl, 0)})
    elif timeframe == "week":
        for i in range(7):
            t = ref_time - timedelta(days=i)
            lbl = t.strftime(group_format)
            timeline_data.insert(0, {"label": lbl, "count": timeline_map.get(lbl, 0)})
    elif timeframe == "quarter":
        for i in range(12):
            t = ref_time - timedelta(weeks=i)
            lbl = t.strftime(group_format)
            timeline_data.insert(0, {"label": lbl, "count": timeline_map.get(lbl, 0)})
    elif timeframe == "year":
        for i in range(12):
            t = ref_time - timedelta(days=i*30)
            lbl = t.strftime(group_format)
            timeline_data.insert(0, {"label": lbl, "count": timeline_map.get(lbl, 0)})

    # 6. Operator Behavior & Performance Analytics
    operators = db.query(Operator).filter(Operator.active == True).all()
    operator_analytics = []
    
    for op in operators:
        op_logs = db.query(OperationLog).filter(OperationLog.operator_id == op.operator_id).all()
        if not op_logs:
            continue
            
        shift_count = len(op_logs)
        total_hours = sum([float(log.hourmeter_end - log.hourmeter_start) for log in op_logs if log.hourmeter_end is not None and log.hourmeter_start is not None])
        
        total_reported_failures = db.query(FailureLog).filter(FailureLog.created_by == op.operator_id).count()
        before_shift_failures = db.query(FailureLog).filter(FailureLog.created_by == op.operator_id, FailureLog.phase == 'before_shift').count()
        during_shift_failures = db.query(FailureLog).filter(FailureLog.created_by == op.operator_id, FailureLog.phase == 'during_shift').count()
        
        safety_violations = sum([1 for log in op_logs if log.work_type == 'production' and log.condition_before_shift == 'broken'])
        
        compliance_score = 100
        compliance_score -= (safety_violations * 20)
        if during_shift_failures > 0 and before_shift_failures == 0:
            compliance_score -= 10
            
        compliance_score = max(0, min(100, compliance_score))
        
        if compliance_score >= 90:
            rank = "Tốt (Tuân thủ cao)"
        elif compliance_score >= 70:
            rank = "Trung bình (Cần nhắc nhở)"
        else:
            rank = "Yếu (Vi phạm quy trình)"
            
        operator_analytics.append({
            "operator_id": op.operator_id,
            "full_name": op.full_name,
            "department": op.department or "Tổ Vận Hành",
            "shift_count": shift_count,
            "total_hours": round(total_hours, 1),
            "reported_failures": total_reported_failures,
            "before_shift_failures": before_shift_failures,
            "during_shift_failures": during_shift_failures,
            "safety_violations": safety_violations,
            "compliance_score": compliance_score,
            "rank": rank
        })
        
    operator_analytics.sort(key=lambda x: (x["compliance_score"], x["shift_count"]), reverse=True)

    return {
        "status_summary": status_summary,
        "timeframe": timeframe,
        "total_failures": len(failures),
        "total_repairs": total_repairs,
        "total_unresolved": total_unresolved,
        "vehicle_freq": vehicle_freq,
        "category_freq": category_freq,
        "timeline_data": timeline_data,
        "operator_analytics": operator_analytics
    }
