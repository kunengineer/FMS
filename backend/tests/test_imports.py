import pytest
import io
import openpyxl
from datetime import datetime, date, time
from app.models import Vehicle, OperationLog, FailureLog, RepairLog, Operator, ChecklistItem

def test_import_activity_success(client, admin_headers, db):
    # Create mock excel file in memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Xe cuốc"
    
    # Rows 1-3: Headers
    ws.append([]) # R1
    ws.append(["STT", "Thời gian", "Phương tiện", None, "Hư hỏng", "Từ", "Đến", "Thời gian hư", "Ghi chú", "Người thực hiện", "GM"]) # R2
    ws.append([None, None, "SK 250", "SK 300", None, None, None, None, None, None, None]) # R3
    
    # R4: row with failure and repair
    ws.append([1, datetime(2026, 5, 2), None, "x", "Hư hỏng", time(8, 0), time(8, 30), "0.5", "Máy lạnh không lạnh", "Phương", 849.0])
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    response = client.post(
        "/api/imports/activity",
        files={"file": ("test_activity.xlsx", file_stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=admin_headers
    )
    assert response.status_code == 200, response.text
    data = response.json()
    assert "statistics" in data
    stats = data["statistics"]
    assert stats["sheets_processed"] == 1
    assert stats["rows_processed"] == 1
    
    # Verify vehicle was created/found
    veh = db.query(Vehicle).filter(Vehicle.vehicle_code == "SK 300").first()
    assert veh is not None
    assert float(veh.current_hourmeter) == 849.0
    
    # Verify failure log was created
    fail = db.query(FailureLog).filter(FailureLog.vehicle_id == veh.vehicle_id).first()
    assert fail is not None
    assert "Máy lạnh không lạnh" in fail.description
    
    # Verify repair log was NOT created
    rep = db.query(RepairLog).filter(RepairLog.failure_id == fail.failure_id).first()
    assert rep is None

def test_import_checklist_success(client, admin_headers, db):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Câu trả lời biểu mẫu 1"
    
    ws.append([
        "Dấu thời gian", "Họ Tên Người Vận Hành", "Tên phương tiện", "Số giờ hoạt động",
        "Kiểm tra mức nhớt động cơ", "Kiểm tra hệ thống phanh chân/tay"
    ])
    ws.append([
        datetime(2026, 5, 2, 7, 30), "Võ Văn Hà", "XC-01", 3095.0,
        "Ok", "Không ăn"
    ])
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    response = client.post(
        "/api/imports/checklist",
        files={"file": ("test_checklist.xlsx", file_stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=admin_headers
    )
    assert response.status_code == 200, response.text
    data = response.json()
    assert "statistics" in data
    stats = data["statistics"]
    assert stats["rows_processed"] == 1
    assert stats["operation_logs_created"] == 1
    
    # Verify checklist items created
    op_log = db.query(OperationLog).filter(OperationLog.hourmeter_start == 3095.0).first()
    assert op_log is not None
    assert op_log.condition_before_shift == "broken"

def test_import_checklist_with_failures_and_safety(client, admin_headers, db):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Câu trả lời biểu mẫu 1"
    
    ws.append([
        "Dấu thời gian", "Họ Tên Người Vận Hành", "Tên phương tiện", "Số giờ hoạt động",
        "Kiểm tra mức nhớt động cơ", "Cần cẩu đảm bảo an toàn để bắt đầu làm việc",
        "Cẩu hư hỏng trong ca làm việc ghi chú hư hỏng đã khắc phục hay chưa? từ mấy giờ đến mấy giờ."
    ])
    # Case: An unsafe shift with failure
    ws.append([
        datetime(2026, 5, 2, 7, 30), "Võ Văn Hà", "LB40-02", 9999.0,
        "Ok", "Không an toàn -yêu cầu ngưng kiêm tra khắc phục sửa chữa", "Bê bạc đąn móc kéo hàng từ 9h"
    ])
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    response = client.post(
        "/api/imports/checklist",
        files={"file": ("test_checklist_fail.xlsx", file_stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=admin_headers
    )
    assert response.status_code == 200, response.text
    data = response.json()
    stats = data["statistics"]
    assert stats["operation_logs_created"] == 1
    assert stats["failure_logs_created"] == 1
    assert stats["repair_logs_created"] == 0
    
    op_log = db.query(OperationLog).filter(OperationLog.hourmeter_start == 9999.0).first()
    assert op_log is not None
    assert op_log.is_safety_confirmed is False
    
    fail = db.query(FailureLog).filter(FailureLog.operation_id == op_log.operation_id).first()
    assert fail is not None
    assert "Bê bạc đąn" in fail.description
    
    rep = db.query(RepairLog).filter(RepairLog.failure_id == fail.failure_id).first()
    assert rep is None

def test_import_weekly_report(client, admin_headers, db):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tuần 23"
    
    # Category divider
    ws.append(["I. NHÓM CẦN CẨU", None, None, None, None, None, None, None, None, None, None])
    
    # Headers
    ws.append([
        "STT", "Tên thiết bị phương tiện", "Giờ hoạt động", "Hoạt động", "Không hoạt động", 
        "Hiện trạng", "Thời gian kiểm tra hư hỏng, sửa chữa trong tuần (h)", "Số lần sửa chữa trong tuần", 
        "Sửa chữa hoàn thành", "Sửa đọng chưa xử lý triệt để", "Ghi chú"
    ])
    
    # Data row 1: broken, unresolved
    ws.append([
        1, "Liebherr CBB 32(40)/32-LB40-1", 49.0, None, "x", 
        "Hệ thống thủy lực xì nhớt", "10h", 1, None, "x", "Chờ vật tư thay thế"
    ])
    
    # Data row 2: normal/active
    ws.append([
        2, "Kobelco SK300", 52.0, "x", None, 
        "Bình thường", None, None, None, None, "Hoạt động ổn định"
    ])
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    response = client.post(
        "/api/imports/weekly-report",
        files={"file": ("test_weekly.xlsx", file_stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=admin_headers
    )
    assert response.status_code == 200, response.text
    data = response.json()
    stats = data["statistics"]
    assert stats["rows_processed"] == 2
    assert stats["failure_logs_created"] == 1
    assert stats["repair_logs_created"] == 0
    
    # Verify vehicle created or updated
    v1 = db.query(Vehicle).filter(Vehicle.vehicle_code == "LB40-1").first()
    assert v1 is not None
    assert v1.current_hourmeter == 49.0
    assert v1.status == "repairing"
    
    from sqlalchemy import func
    v2 = db.query(Vehicle).filter(func.replace(Vehicle.vehicle_code, ' ', '') == "SK300").first()
    assert v2 is not None
    assert float(v2.current_hourmeter) == 901.0
    assert v2.status == "active"
    
    # Verify failure log and unresolved status
    fail = db.query(FailureLog).filter(FailureLog.vehicle_id == v1.vehicle_id).first()
    assert fail is not None
    assert "Hệ thống thủy lực xì nhớt" in fail.description
    assert fail.is_repaired is False
    
    # Verify RepairLog was NOT created
    rep = db.query(RepairLog).filter(RepairLog.failure_id == fail.failure_id).first()
    assert rep is None

def test_import_weekly_report_robust(client, admin_headers, db):
    # 1. Create a sheet with a non-tuan name but valid headers (e.g. Sheet1)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MẪU_BÁO_CÁO"
    
    # Report info
    ws.append(["Mẫu Báo Cáo Tuần Tiêu Chuẩn"])
    ws.append(["Số Tuần:", "Tuần 24"])
    ws.append(["Ngày Báo Cáo:", "15/07/2026"])
    ws.append([])
    
    # Headers matching layout B/C
    ws.append([
        "STT", "Mã PT", "Tên Thiết Bị", "Thông số KT", "Giờ đầu", "Giờ cuối", "Giờ HĐ", 
        "Tình trạng", "Chi Tiết Hư Hỏng / Hiện Trạng", "Thời Gian Dừng (h)", "Khắc phục?", "Ghi chú"
    ])
    
    # Data row: active vehicle with failure under monitoring
    ws.append([
        1, "PC200-7", "Xe cuốc Komatsu PC200", "145HP", 512.0, 528.0, 16.0,
        "Bình thường", "Đã xử lý xì nhớt gầm, đang theo dõi", 4.0, "Không", "Theo dõi sát"
    ])
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    response = client.post(
        "/api/imports/weekly-report",
        files={"file": ("test_weekly_robust.xlsx", file_stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=admin_headers
    )
    assert response.status_code == 200, response.text
    
    # Verify vehicle hourmeters are imported properly
    v = db.query(Vehicle).filter(Vehicle.vehicle_code == "PC200-7").first()
    assert v is not None
    # Hourmeter should be 528.0 (from Giờ cuối)
    assert float(v.current_hourmeter) == 528.0
    
    # Verify OperationLog hourmeters
    op_log = db.query(OperationLog).filter(OperationLog.vehicle_id == v.vehicle_id).first()
    assert op_log is not None
    assert float(op_log.hourmeter_start) == 512.0
    assert float(op_log.hourmeter_end) == 528.0
    
    # Operator must be the currently logged-in user (admin)
    # The default mock user in admin_headers has username 'admin' or ID ME_ADMIN/OP_ADMIN depending on seeds.
    # Let's verify that the operator is created/mapped
    assert op_log.operator_id is not None
    
    # Verify FailureLog is under monitoring (is_repaired = False)
    fail = db.query(FailureLog).filter(FailureLog.vehicle_id == v.vehicle_id).first()
    assert fail is not None
    assert "Đã xử lý xì nhớt gầm, đang theo dõi" in fail.description
    assert fail.is_repaired is False
    
    # Verify RepairLog was NOT created
    rep = db.query(RepairLog).filter(RepairLog.failure_id == fail.failure_id).first()
    assert rep is None

def test_import_weekly_report_iso_week(client, admin_headers, db):
    # Create workbook with sheet name in YYYY-WW format (e.g. 2026-W12)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026-W12"
    
    ws.append(["Mẫu Báo Cáo Tuần 2026-W12"])
    ws.append([])
    ws.append([])
    ws.append([])
    
    ws.append([
        "STT", "Tên thiết bị phương tiện", "Mã PT", "Giờ hoạt động", "Hiện trạng", "Ghi chú"
    ])
    
    ws.append([
        1, "Xe nâng Hangcha 3T", "HC30-1", 24.5, "Hoạt động ổn định", "Không có hư hỏng"
    ])
    
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    response = client.post(
        "/api/imports/weekly-report",
        files={"file": ("test_weekly_iso.xlsx", file_stream, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=admin_headers
    )
    assert response.status_code == 200, response.text
    
    # Assert Friday of 2026-W12 is parsed correctly:
    # 2026 ISO week 1 starts on Mon Jan 29, 2025? No, Jan 4 is a Sunday in 2026.
    # W12 Friday should be parsed. Let's verify we have the log.
    v = db.query(Vehicle).filter(Vehicle.vehicle_code == "HC30-1").first()
    assert v is not None
    
    op_log = db.query(OperationLog).filter(OperationLog.vehicle_id == v.vehicle_id).first()
    assert op_log is not None
    # Let's verify the calculated work_date is in March 2026 (week 12 Friday)
    assert op_log.work_date.year == 2026
    assert op_log.work_date.month == 3
    # Week 12 Friday of 2026 is March 20, 2026
    assert op_log.work_date.day == 20


