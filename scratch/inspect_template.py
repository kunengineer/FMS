import openpyxl
import os
import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "backend")))
sys.path.append(os.path.abspath("."))

# Configure stdout
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

filepath = "d:/Project/Fleet-Manager-System/Mau_Bao_Cao_Chuan.xlsx"
if not os.path.exists(filepath):
    print("File not found.")
    sys.exit(1)

wb = openpyxl.load_workbook(filepath, data_only=True)
print(f"Sheet names in Mau_Bao_Cao_Chuan.xlsx: {wb.sheetnames}")

from app.routers.imports import extract_date_from_sheet, find_header_row_and_cols
from app.routers.imports import ExcelWrapper

with open(filepath, "rb") as f:
    file_bytes = f.read()
wrapper = ExcelWrapper(file_bytes, filename=filepath)

for name in wb.sheetnames:
    sheet = wrapper.get_sheet(name)
    work_date = extract_date_from_sheet(sheet, wrapper, sheet_name=name, filename="Mau_Bao_Cao_Chuan.xlsx")
    header_row_idx, cols = find_header_row_and_cols(sheet, wrapper)
    print(f"\nSheet name: {name}")
    print(f"  Parsed work_date: {work_date}")
    print(f"  Header row index: {header_row_idx}")
    print(f"  Columns mapping: {cols}")
    if header_row_idx:
        max_rows = wrapper.get_max_row(sheet)
        non_empty = 0
        for r in range(header_row_idx + 1, max_rows + 1):
            val = wrapper.get_cell_value(sheet, r, cols["name"])
            if val is not None and str(val).strip() != "":
                non_empty += 1
        print(f"  Non-empty rows after header: {non_empty}")
